from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from flask_pymongo import PyMongo
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import os
from functools import wraps
from theme_config import get_active_fe_theme, get_active_noc_theme, get_theme_for_role

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MONGO_URI'] = os.environ.get('MONGO_URI', 'mongodb://localhost:27017/sdwan_tracker')
#prod
# app.config['MONGO_URI'] = os.environ.get('MONGO_URI', 'mongodb://myAdminUser:MyStrongPassword123@localhost:27017/sdwan_tracker?authSource=admin')

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

mongo = PyMongo(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# ─── Jinja2 template filters ───────────────────────────────────────────────
@app.template_filter('display_name')
def display_name_filter(username):
    """Convert 'Name_Franchise' to 'Name (Franchise)'."""
    if not username:
        return username
    if '_' in username:
        parts = username.split('_', 1)
        return f'{parts[0]} ({parts[1]})'
    return username

# ─── Canonical role constants ───────────────────────────────────────────────
# These are the ONLY valid role strings used throughout the application.
# Session, DB documents, and route guards all use these exact strings.
ROLE_FE             = 'FIELD_ENGINEER'
ROLE_FEG            = 'FIELD_ENGINEER_GROUP'
ROLE_FS             = 'FIELD_SUPPORT'
ROLE_FSG            = 'FIELD_SUPPORT_GROUP'
ROLE_NS             = 'NOC_SUPPORT'      # NOC Support individual operator
ROLE_NSG            = 'NOC_SUPPORT_GROUP'
ROLE_ANALYTICS      = 'ANALYTICS'

FE_ROLES  = {ROLE_FE, ROLE_FEG, ROLE_FS, ROLE_FSG}
NOC_ROLES = {ROLE_NS, ROLE_NSG}

# ─── Real-time configuration ─────────────────────────────────────────────────
# Controls how updates are delivered: Socket.IO-first with API fallback
REALTIME_MODE = os.environ.get('REALTIME_MODE', 'hybrid')  # 'socket', 'api', 'hybrid'
SOCKET_TIMEOUT = int(os.environ.get('SOCKET_TIMEOUT', '2000'))  # milliseconds
SOCKET_INCLUDE_FULL_DATA = os.environ.get('SOCKET_INCLUDE_FULL_DATA', 'true').lower() == 'true'

# ─── Tracker status constants ────────────────────────────────────────────────
# Canonical status strings used across backend, frontend, and analytics.
STATUS_WAITING_NOC       = 'waiting_noc_assignment'
STATUS_NOC_WORKING       = 'noc_working'
STATUS_ZTP_PULL_PENDING  = 'ztp_pull_pending'
STATUS_ZTP_CONFIG_UNVERIFIED = 'ztp_config_unverified'
STATUS_ZTP_PULL_DONE_FE  = 'ztp_pull_done_by_fe'
STATUS_ZTP_PULL_UNVERIFIED = 'ztp_pull_unverified'
STATUS_ZTP_PULL_REQ_NOC  = 'ztp_pull_requested_from_noc'
STATUS_FE_REQ_ZTP        = 'fe_requested_ztp'
STATUS_READY_COORD       = 'ready_for_coordination'
STATUS_HSO_SUBMITTED     = 'hso_submitted'
STATUS_HSO_REJECTED      = 'hso_rejected'
STATUS_COMPLETE          = 'installation_complete'
# Legacy statuses (kept for backwards compat with old documents)
STATUS_ZTP_PULL_VERIFIED = 'ztp_pull_verified'
STATUS_ZTP_PULL_DONE_NOC = 'ztp_pull_done_by_noc'

# Statuses where chat is unlocked (FE-NS coordination phase)
CHAT_UNLOCKED_STATUSES = {
    STATUS_READY_COORD,
    STATUS_FE_REQ_ZTP,
    STATUS_ZTP_PULL_REQ_NOC,    # Chat unlocks when FE requests NOC to do ZTP pull
    STATUS_HSO_SUBMITTED,
    STATUS_HSO_REJECTED,
    STATUS_COMPLETE,
    STATUS_ZTP_PULL_VERIFIED,   # Legacy
    STATUS_ZTP_PULL_DONE_NOC,   # Legacy
}

# Statuses from which FE can submit HSO
HSO_SUBMITTABLE_STATUSES = {
    STATUS_READY_COORD,
    STATUS_HSO_REJECTED,
    STATUS_FE_REQ_ZTP,
    STATUS_ZTP_PULL_VERIFIED,   # Legacy
    STATUS_ZTP_PULL_DONE_NOC,   # Legacy
}

# ─── Helpers ────────────────────────────────────────────────────────────────
@app.context_processor
def inject_config():
    """Make configuration available to all templates"""
    return {
        'config': {
            'REALTIME_MODE': REALTIME_MODE,
            'SOCKET_TIMEOUT': SOCKET_TIMEOUT,
            'SOCKET_INCLUDE_FULL_DATA': SOCKET_INCLUDE_FULL_DATA
        }
    }

def get_utc_now():
    """Return the current time as a naive UTC datetime.
    MongoDB stores datetimes as UTC by default; using naive datetimes
    avoids timezone-aware vs naive comparison errors."""
    return datetime.utcnow()


def serialize_doc(doc):
    """Recursively convert ObjectIds and datetimes to JSON-safe types.
    Datetimes are formatted as ISO-8601 with a 'Z' suffix so the
    frontend knows they are UTC and can convert to IST."""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    if isinstance(doc, dict):
        doc = doc.copy()
        if '_id' in doc:
            doc['_id'] = str(doc['_id'])
        for key, value in doc.items():
            if isinstance(value, ObjectId):
                doc[key] = str(value)
            elif isinstance(value, datetime):
                doc[key] = value.isoformat() + 'Z' if value.tzinfo is None else value.isoformat()
            elif isinstance(value, (dict, list)):
                doc[key] = serialize_doc(value)
        return doc
    return doc


def make_event(stage, actor_id, actor_role, remarks, metadata=None):
    """Build a standard event dict. Every stage transition goes through
    this helper to ensure consistent structure across all API endpoints."""
    return {
        'stage': stage,
        'timestamp': get_utc_now(),
        'actor': actor_id,
        'actor_role': actor_role,
        'remarks': remarks,
        'metadata': metadata or {},
        'delay_tags': []
    }


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def is_chat_unlocked(tracker):
    """Chat unlocks when the tracker enters the FE-NS coordination phase.
    This happens via multiple paths:
    1. NS clicked 'Mark Ready for Coordination' (normal happy path).
    2. FE clicked 'Request NS to do ZTP' (legacy flow).
    3. FE clicked 'Request NOC to Perform ZTP Pull' (new 2-phase flow).
    4. HSO submitted/rejected — chat stays open for discussion.
    5. Installation complete — chat available for post-completion review."""
    return tracker.get('status') in CHAT_UNLOCKED_STATUSES


# ─── Page Routes ────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' in session:
        role = session.get('role')
        if role in FE_ROLES:
            return redirect(url_for('fe_dashboard'))
        elif role in NOC_ROLES:
            return redirect(url_for('noc_dashboard'))
        elif role == ROLE_ANALYTICS:
            return redirect(url_for('analytics_dashboard'))
    return redirect(url_for('login'))


@app.route('/login')
def login():
    return render_template('login.html')


@app.route('/fe/dashboard')
@login_required
def fe_dashboard():
    if session.get('role') not in FE_ROLES:
        return redirect(url_for('index'))
    return render_template('fe_dashboard.html', theme=get_theme_for_role(session.get('role')))


@app.route('/fe/new-installation')
@login_required
def fe_new_installation():
    if session.get('role') != ROLE_FE:
        return redirect(url_for('fe_dashboard'))
    return render_template('fe_new_installation.html', theme=get_theme_for_role(session.get('role')))


@app.route('/fe/tracker/<tracker_id>')
@login_required
def fe_tracker_detail(tracker_id):
    if session.get('role') not in FE_ROLES:
        return redirect(url_for('index'))
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        flash('Tracker not found', 'error')
        return redirect(url_for('fe_dashboard'))
    # Only the FE who created the tracker can take actions (FEG/FS/FSG are read-only viewers)
    can_interact = (tracker.get('fe', {}).get('id') == session['user_id'])
    return render_template('fe_tracker_detail.html',
                           tracker_id=tracker_id,
                           can_interact=can_interact,
                           theme=get_theme_for_role(session.get('role')))


@app.route('/noc/dashboard')
@login_required
def noc_dashboard():
    if session.get('role') not in NOC_ROLES:
        return redirect(url_for('index'))
    return render_template('noc_dashboard.html', theme=get_theme_for_role(session.get('role')))


@app.route('/noc/tracker/<tracker_id>')
@login_required
def noc_tracker_detail(tracker_id):
    if session.get('role') not in NOC_ROLES:
        return redirect(url_for('index'))
    return render_template('noc_tracker_detail.html',
                           tracker_id=tracker_id,
                           theme=get_theme_for_role(session.get('role')))


@app.route('/analytics/dashboard')
@login_required
def analytics_dashboard():
    if session.get('role') not in {ROLE_ANALYTICS, ROLE_NSG, ROLE_FSG}:
        return redirect(url_for('index'))
    return render_template('analytics_dashboard_v1.html')


# Legacy redirects so old bookmarks still work
@app.route('/franchise/dashboard')
@login_required
def franchise_dashboard():
    return redirect(url_for('fe_dashboard'))


@app.route('/field_support/dashboard')
@login_required
def field_support_dashboard():
    return redirect(url_for('fe_dashboard'))


@app.route('/field_support_admin/dashboard')
@login_required
def field_support_admin_dashboard():
    return redirect(url_for('fe_dashboard'))


# ─── Login / Logout API ─────────────────────────────────────────────────────
@app.route('/api/login/field-support-groups', methods=['GET'])
def api_get_field_support_groups():
    # Query users collection for FIELD_SUPPORT_GROUP role
    groups = list(mongo.db.users.find(
        {'role': ROLE_FSG},
        {'_id': 0, 'username': 1, 'name': 1, 'zone': 1, 'role': 1}
    ))
    return jsonify(groups)


@app.route('/api/login/field-supports', methods=['GET'])
def api_get_field_supports():
    # Query users collection for FIELD_SUPPORT role only (FSG is fetched separately)
    supports = list(mongo.db.users.find(
        {'role': ROLE_FS},
        {'_id': 0, 'username': 1, 'name': 1, 'region': 1, 'role': 1}
    ))
    return jsonify(supports)


@app.route('/api/login/field-engineer-groups', methods=['GET'])
def api_get_field_engineer_groups():
    # Query users collection for FIELD_ENGINEER_GROUP role
    groups = list(mongo.db.users.find(
        {'role': ROLE_FEG},
        {'_id': 0, 'username': 1, 'name': 1, 'region': 1, 'role': 1}
    ))
    return jsonify(groups)


@app.route('/api/login/field-engineers', methods=['GET'])
def api_get_field_engineers():
    feg_name = request.args.get('field_engineer_group')
    query = {'role': ROLE_FE}
    if feg_name:
        query['field_engineer_group'] = feg_name
    engineers = list(mongo.db.users.find(
        query,
        {'_id': 0, 'username': 1, 'name': 1, 'field_engineer_group': 1, 'region': 1, 'location': 1, 'role': 1}
    ))
    return jsonify(engineers)


@app.route('/api/login/noc-supports', methods=['GET'])
def api_get_noc_supports():
    noc_users = list(mongo.db.users.find(
        {'role': {'$in': [ROLE_NS, ROLE_NSG]}},
        {'_id': 0, 'username': 1, 'name': 1, 'role': 1}
    ))
    return jsonify(noc_users)


# ─── Get NOC Users for Reassignment ─────────────────────────────────────────
@app.route('/api/noc-users', methods=['GET'])
@login_required
def api_get_noc_users_for_reassignment():
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get all NOC users except current user
    noc_users = list(mongo.db.users.find(
        {
            'role': ROLE_NS,
            '_id': {'$ne': ObjectId(session['user_id'])}
        },
        {'_id': 1, 'username': 1, 'name': 1}
    ))
    
    # Convert ObjectId to string
    for user in noc_users:
        user['_id'] = str(user['_id'])
    
    return jsonify({'noc_users': noc_users})


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.json
    role = data.get('role')
    password = data.get('password')

    query = {'role': role}

    if role == ROLE_FE:
        query['name'] = data.get('fe_name')
        query['field_engineer_group'] = data.get('fe_group')
    elif role == ROLE_FEG:
        query['name'] = data.get('feg_name')
        # Removed state query - FEG no longer has state field
    elif role == ROLE_FS:
        query['name'] = data.get('fs_name')
    elif role == ROLE_FSG:
        query['name'] = data.get('fsg_name')
    elif role in NOC_ROLES:
        query['username'] = data.get('noc_username')
    elif role == ROLE_ANALYTICS:
        query['username'] = data.get('username', 'analytics')
    else:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400

    user = mongo.db.users.find_one(query)
    if not user:
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    # Verify password using Werkzeug's check_password_hash for hashed passwords
    if not check_password_hash(user.get('password', ''), password):
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

    session['user_id'] = str(user['_id'])
    session['username'] = user.get('username', user.get('name'))
    session['role'] = role
    session['name'] = user.get('name', user.get('username'))

    if role == ROLE_FE:
        session['field_engineer_group'] = user.get('field_engineer_group')
        session['field_support'] = user.get('field_support')
        session['region'] = user.get('region')
        session['zone'] = user.get('zone')
        session['email'] = user.get('email')
        session['contact'] = user.get('contact')
        session['location'] = user.get('location')
    elif role == ROLE_FEG:
        session['field_engineer_group'] = user.get('name')
        session['field_support'] = user.get('field_support')
        session['state'] = user.get('state')
        session['region'] = user.get('region')
        session['zone'] = user.get('zone')
    elif role == ROLE_FS:
        session['field_support'] = user.get('name')
        session['region'] = user.get('region')
        session['field_support_group'] = user.get('field_support_group')
        session['zone'] = user.get('zone')
    elif role == ROLE_FSG:
        session['field_support_group'] = user.get('name')
        session['zone'] = user.get('zone')
    elif role in NOC_ROLES:
        session['noc_name'] = user.get('name', user.get('username'))

    return jsonify({'success': True, 'role': role})


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/debug/time')
def api_debug_time():
    import time
    now_utc = get_utc_now()
    now_local = datetime.now()
    return jsonify({
        'server_time_utc': now_utc.isoformat(),
        'server_time_local': now_local.isoformat(),
        'system_timezone_offset_seconds': time.timezone,
        'note': 'All times stored in UTC; convert to IST in frontend (+5:30)'
    })


# ─── Tracker Query APIs ─────────────────────────────────────────────────────
@app.route('/api/trackers/check/<sdwan_id>')
@login_required
def api_check_sdwan_id(sdwan_id):
    tracker = mongo.db.trackers.find_one({'sdwan_id': sdwan_id})
    if tracker:
        return jsonify({'exists': True, 'tracker': serialize_doc(tracker)})
    return jsonify({'exists': False})


@app.route('/api/trackers/<tracker_id>')
@login_required
def api_get_tracker(tracker_id):
    import time
    start_time = time.time()
    
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    query_time = time.time() - start_time
    
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    
    can_interact = True
    if session.get('role') == ROLE_FE:
        can_interact = (tracker.get('fe', {}).get('id') == session['user_id'])
    
    serialize_start = time.time()
    result = {'tracker': serialize_doc(tracker), 'can_interact': can_interact}
    serialize_time = time.time() - serialize_start
    total_time = time.time() - start_time
    
    print(f"[API Performance] GET /api/trackers/{tracker_id}: Query={query_time*1000:.1f}ms, Serialize={serialize_time*1000:.1f}ms, Total={total_time*1000:.1f}ms")
    
    return jsonify(result)


@app.route('/api/trackers/all-fe')
@login_required
def api_all_fe():
    """Return trackers visible to the current FE-side user based on hierarchy."""
    role = session.get('role')
    if role == ROLE_FE:
        trackers = list(mongo.db.trackers.find({'fe.id': session['user_id']}).sort('created_at', -1))
    elif role == ROLE_FEG:
        feg = session.get('field_engineer_group')
        trackers = list(mongo.db.trackers.find({'fe.field_engineer_group': feg}).sort('created_at', -1))
    elif role == ROLE_FS:
        # FS sees all trackers from FEGs under their region
        fs = session.get('field_support')
        # Get all FEG names under this FS
        feg_users = list(mongo.db.users.find({'role': ROLE_FEG, 'field_support': fs}, {'name': 1}))
        fegs = [feg['name'] for feg in feg_users if 'name' in feg]
        if fegs:
            trackers = list(mongo.db.trackers.find({'fe.field_engineer_group': {'$in': fegs}}).sort('created_at', -1))
        else:
            trackers = []
    elif role == ROLE_FSG:
        trackers = list(mongo.db.trackers.find().sort('created_at', -1))
    else:
        trackers = []
    return jsonify({'trackers': serialize_doc(trackers)})


@app.route('/api/trackers/all-noc')
@login_required
def api_all_noc_trackers():
    if session.get('role') not in NOC_ROLES:
        return jsonify({'error': 'Unauthorized'}), 403
    trackers = list(mongo.db.trackers.find().sort('created_at', -1))
    return jsonify({'trackers': serialize_doc(trackers)})


@app.route('/api/trackers/unassigned')
@login_required
def api_unassigned_trackers():
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403
    trackers = list(
        mongo.db.trackers.find({'noc_assignee': None, 'status': 'waiting_noc_assignment'}).sort('created_at', 1)
    )
    return jsonify({'trackers': serialize_doc(trackers)})


@app.route('/api/trackers/my-installations')
@login_required
def api_my_installations():
    role = session.get('role')
    if role == ROLE_FE:
        trackers = list(mongo.db.trackers.find({'fe.id': session['user_id']}).sort('created_at', -1).limit(50))
    elif role == ROLE_NS:
        trackers = list(mongo.db.trackers.find({'noc_assignee': session['user_id']}).sort('created_at', -1))
    else:
        trackers = []
    return jsonify({'trackers': serialize_doc(trackers)})


@app.route('/api/noc/users/stats')
@login_required
def api_noc_users_stats():
    if session.get('role') != ROLE_NSG:
        return jsonify({'error': 'Unauthorized'}), 403
    noc_users = list(mongo.db.users.find({'role': ROLE_NS}))
    users_stats = []
    for user in noc_users:
        user_id = str(user['_id'])
        ongoing_count = mongo.db.trackers.count_documents({
            'noc_assignee': user_id,
            'status': {'$ne': 'installation_complete'}
        })
        completed_count = mongo.db.trackers.count_documents({
            'noc_assignee': user_id,
            'status': 'installation_complete'
        })
        users_stats.append({
            'id': user_id,
            'name': user.get('name', 'N/A'),
            'username': user.get('username', 'N/A'),
            'ongoing_count': ongoing_count,
            'completed_count': completed_count,
            'total_count': ongoing_count + completed_count
        })
    return jsonify({'users': users_stats})


@app.route('/api/hierarchy/view')
@login_required
def api_hierarchy_view():
    role = session.get('role')
    
    # Get date range parameters (optional)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Build date filters if provided
    created_date_filter = {}
    completed_date_filter = {}
    if date_from and date_to:
        try:
            from_dt = datetime.fromisoformat(date_from.replace('Z', ''))
            to_dt = datetime.fromisoformat(date_to.replace('Z', ''))
            created_date_filter = {'created_at': {'$gte': from_dt, '$lte': to_dt}}
            completed_date_filter = {'completed_at': {'$gte': from_dt, '$lte': to_dt}}
        except:
            pass  # If date parsing fails, ignore date filter

    if role == ROLE_FEG:
        feg = session.get('field_engineer_group')
        fes = list(mongo.db.users.find({'role': ROLE_FE, 'field_engineer_group': feg}))
        data = []
        for fe in fes:
            fe_id = str(fe['_id'])
            base_filter = {'fe.id': fe_id}

            pending = mongo.db.trackers.count_documents({**base_filter, 'status': 'waiting_noc_assignment', **created_date_filter})
            ongoing = mongo.db.trackers.count_documents({**base_filter, 'status': {'$nin': ['waiting_noc_assignment', 'installation_complete']}, **created_date_filter})
            done    = mongo.db.trackers.count_documents({**base_filter, 'status': 'installation_complete', **completed_date_filter})
            total   = pending + ongoing + done
            
            data.append({
                'id': fe_id, 
                'name': fe.get('name'), 
                'phone': fe.get('contact'),  # Use 'contact' field from user document
                'email': fe.get('email'),
                'location': fe.get('location'),  # State/location
                'field_support': session.get('field_support'),
                'total_count': total, 
                'unassigned_count': pending, 
                'ongoing_count': ongoing, 
                'completed_count': done
            })
        return jsonify({'type': 'field_engineers', 'data': data, 'total_count': len(data)})

    elif role == ROLE_FS:
        fs = session.get('field_support')
        fegs = list(mongo.db.users.find({'role': ROLE_FEG, 'field_support': fs}))
        data = []
        for feg in fegs:
            feg_name = feg.get('name')
            base_filter = {'fe.field_engineer_group': feg_name}

            pending = mongo.db.trackers.count_documents({**base_filter, 'status': 'waiting_noc_assignment', **created_date_filter})
            ongoing = mongo.db.trackers.count_documents({**base_filter, 'status': {'$nin': ['waiting_noc_assignment', 'installation_complete']}, **created_date_filter})
            done    = mongo.db.trackers.count_documents({**base_filter, 'status': 'installation_complete', **completed_date_filter})
            total   = pending + ongoing + done
            
            fe_count = mongo.db.users.count_documents({'role': ROLE_FE, 'field_engineer_group': feg_name})
            data.append({
                'id': str(feg['_id']), 
                'name': feg_name, 
                'region': feg.get('region'),  # Add region field
                'fe_count': fe_count,
                'total_count': total, 
                'unassigned_count': pending, 
                'ongoing_count': ongoing, 
                'completed_count': done
            })
        return jsonify({'type': 'field_engineer_groups', 'data': data, 'total_count': len(data)})

    elif role == ROLE_FSG:
        fss = list(mongo.db.users.find({'role': ROLE_FS}))
        data = []
        for fs in fss:
            fs_name = fs.get('name')
            
            # Get all FEGs under this FS
            fegs_under_fs = list(mongo.db.users.find({'role': ROLE_FEG, 'field_support': fs_name}))
            feg_names = [feg.get('name') for feg in fegs_under_fs]
            
            if feg_names:
                # Build base filter with FEG names
                base_filter = {'fe.field_engineer_group': {'$in': feg_names}}

                pending = mongo.db.trackers.count_documents({**base_filter, 'status': 'waiting_noc_assignment', **created_date_filter})
                ongoing = mongo.db.trackers.count_documents({**base_filter, 'status': {'$nin': ['waiting_noc_assignment', 'installation_complete']}, **created_date_filter})
                done    = mongo.db.trackers.count_documents({**base_filter, 'status': 'installation_complete', **completed_date_filter})
                total   = pending + ongoing + done

                fe_count = mongo.db.users.count_documents({'role': ROLE_FE, 'field_engineer_group': {'$in': feg_names}})
            else:
                total = pending = ongoing = done = fe_count = 0
            
            feg_count = len(feg_names)
            
            data.append({'id': str(fs['_id']), 'name': fs_name, 'feg_count': feg_count, 'fe_count': fe_count,
                         'total_count': total, 'unassigned_count': pending, 'ongoing_count': ongoing, 'completed_count': done})
        return jsonify({'type': 'field_supports', 'data': data, 'total_count': len(data)})

    return jsonify({'error': 'Unauthorized'}), 403


@app.route('/api/hierarchy/drill-down')
@login_required
def api_hierarchy_drill_down():
    """Drill-down API for hierarchical views - returns FEGs under an FS, or FEs under a FEG"""
    role = session.get('role')
    
    # Get drill-down parameters
    fs_name = request.args.get('fs_name')
    feg_name = request.args.get('feg_name')
    
    # Get date range parameters (optional)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Build date filters if provided
    created_date_filter = {}
    completed_date_filter = {}
    if date_from and date_to:
        try:
            from_dt = datetime.fromisoformat(date_from.replace('Z', ''))
            to_dt = datetime.fromisoformat(date_to.replace('Z', ''))
            created_date_filter = {'created_at': {'$gte': from_dt, '$lte': to_dt}}
            completed_date_filter = {'completed_at': {'$gte': from_dt, '$lte': to_dt}}
        except:
            pass  # If date parsing fails, ignore date filter
    
    # FSG can drill down to FEGs under an FS
    if role == ROLE_FSG and fs_name:
        fegs = list(mongo.db.users.find({'role': ROLE_FEG, 'field_support': fs_name}))
        data = []
        for feg in fegs:
            feg_name_val = feg.get('name')
            base_filter = {'fe.field_engineer_group': feg_name_val}

            pending = mongo.db.trackers.count_documents({**base_filter, 'status': 'waiting_noc_assignment', **created_date_filter})
            ongoing = mongo.db.trackers.count_documents({**base_filter, 'status': {'$nin': ['waiting_noc_assignment', 'installation_complete']}, **created_date_filter})
            done    = mongo.db.trackers.count_documents({**base_filter, 'status': 'installation_complete', **completed_date_filter})
            total   = pending + ongoing + done
            
            fe_count = mongo.db.users.count_documents({'role': ROLE_FE, 'field_engineer_group': feg_name_val})
            data.append({
                'id': str(feg['_id']), 
                'name': feg_name_val, 
                'phone': feg.get('phone'),
                'email': feg.get('email'),
                'region': feg.get('region'),
                'field_support': feg.get('field_support'),
                'fe_count': fe_count,
                'total_count': total, 
                'unassigned_count': pending, 
                'ongoing_count': ongoing, 
                'completed_count': done
            })
        return jsonify({'success': True, 'type': 'field_engineer_groups', 'data': data, 'total_count': len(data)})
    
    # FSG or FS can drill down to FEs under a FEG
    if (role == ROLE_FSG or role == ROLE_FS) and feg_name:
        fes = list(mongo.db.users.find({'role': ROLE_FE, 'field_engineer_group': feg_name}))
        data = []
        for fe in fes:
            fe_username = fe.get('username')
            # Match trackers by username (stable; stored on every tracker from session['username'])
            base_filter = {'fe.username': fe_username}

            pending = mongo.db.trackers.count_documents({**base_filter, 'status': 'waiting_noc_assignment', **created_date_filter})
            ongoing = mongo.db.trackers.count_documents({**base_filter, 'status': {'$nin': ['waiting_noc_assignment', 'installation_complete']}, **created_date_filter})
            done    = mongo.db.trackers.count_documents({**base_filter, 'status': 'installation_complete', **completed_date_filter})
            total   = pending + ongoing + done

            data.append({
                'id': str(fe['_id']),
                'name': fe.get('name'),
                'phone': fe.get('contact') or fe.get('phone'),
                'email': fe.get('email'),
                'region': fe.get('region'),
                'field_engineer_group': fe.get('field_engineer_group'),
                'total_count': total,
                'unassigned_count': pending,
                'ongoing_count': ongoing,
                'completed_count': done
            })
        return jsonify({'success': True, 'type': 'field_engineers', 'data': data, 'total_count': len(data)})
    
    return jsonify({'success': False, 'error': 'Invalid drill-down parameters'}), 400


# ─── Tracker Creation ───────────────────────────────────────────────────────
@app.route('/api/trackers', methods=['POST'])
@login_required
def api_create_tracker():
    if session.get('role') != ROLE_FE:
        return jsonify({'success': False, 'message': 'Only Field Engineers can create trackers'}), 403

    data = request.json
    existing = mongo.db.trackers.find_one({'sdwan_id': data['sdwan_id']})
    if existing:
        return jsonify({'success': False, 'message': 'SDWAN ID already exists',
                        'tracker_id': str(existing['_id'])}), 409

    now = get_utc_now()
    captured_images = data.get('images', {})

    def build_sim(provider_key, number_key):
        provider = data.get(provider_key, '')
        number   = data.get(number_key, '')
        images   = [
            {'field': 'provider', 'data': captured_images[provider_key]} if captured_images.get(provider_key) else None,
            {'field': 'number',   'data': captured_images[number_key]}   if captured_images.get(number_key)   else None,
        ]
        return {
            'provider': provider,
            'number': number,
            'status': 'pending' if number else 'not_required',
            'images': [i for i in images if i],
            # Per-attempt history — each attempt adds an entry here so retries are traceable
            'attempts': [],
            'failure_reason': None,
            'root_cause_of_initial_failure': None,
            'activation_started_at': None,
        }

    tracker = {
        # ── Identity ──────────────────────────────────────────────────────
        'tracker_id': f"SDWAN-{now.year}-{mongo.db.trackers.count_documents({}) + 1:06d}",
        'sdwan_id': data['sdwan_id'],
        'customer': data['customer'],
        'site_name': data.get('site_name', ''),
        'site_address': data.get('site_address', ''),

        # ── Field Engineer ────────────────────────────────────────────────
        'fe': {
            'id': session['user_id'],
            'username': session['username'],
            'name': session.get('name', session['username']),
            'phone': data['fe_phone'],
            'email': session.get('email', ''),
            # Hierarchy codes for filtering at every FEG/FS/FSG level
            'field_engineer_group': session.get('field_engineer_group'),
            'field_support': session.get('field_support'),
            'region': session.get('region'),
            'zone': session.get('zone'),
            # History allows future multi-FE handoff without losing audit trail
            'history': [{'fe_name': session.get('name', session['username']), 'assigned_at': now, 'left_at': None}]
        },

        # ── NOC Assignment ────────────────────────────────────────────────
        # noc_assignee holds the CURRENT assignee user_id (string or None).
        # noc_history is the full audit log of who worked on this tracker.
        'noc_assignee': None,
        'noc_history': [],

        # ── SIM Data ──────────────────────────────────────────────────────
        'sim': {
            'sim1': build_sim('sim1_provider', 'sim1_number'),
            'sim2': build_sim('sim2_provider', 'sim2_number'),
        },

        # ── Router Details ────────────────────────────────────────────
        'router': {
            'type': data.get('router_type', ''),
            'make': data.get('router_make', ''),
            'firmware_version': data.get('router_firmware_version', ''),
            'images': [{'field': 'firmware', 'data': captured_images['router_firmware_version']}]
                      if captured_images.get('router_firmware_version') else []
        },

        # ── Firmware + ZTP ────────────────────────────────────────────────
        'firmware': {
            'version': data.get('router_firmware_version', ''),
            'images': [{'field': 'version', 'data': captured_images['router_firmware_version']}]
                      if captured_images.get('router_firmware_version') else []
        },
        'ztp': {
            # ztp_config_status: NS verifies that ZTP config is correct
            # values: pending → config_verified | config_failed
            'config_status': 'pending',
            'config_verified_at': None,
            'config_failure_reason': None,
            # ztp_execution: who runs ZTP and the result
            # values: pending → initiated → completed | failed
            # performed_by: 'FE' or 'NS'
            'status': 'pending',
            'performed_by': None,         # 'FE' or 'NS' — key for KPI "ZTP done by FE vs NS"
            'fe_requested_ns': False,     # True when FE cannot do ZTP and asks NS
            'initiated_at': None,
            'completed_at': None,
            'failure_reason': None,
            'root_cause_of_initial_failure': None,
            'attempts': [],               # [{performed_by, started_at, result, reason}]
        },

        # ── HSO ───────────────────────────────────────────────────────────
        # HSO tracks the hand-over sign-off cycle. FE submits, NS approves or rejects.
        # Multiple submit/reject cycles are captured in hso_attempts.
        'hso': {
            'status': 'pending',          # pending | submitted | rejected | approved
            'submitted_at': None,         # Set when FE clicks Submit HSO
            'approved_at': None,          # Set when NS approves
            'rejected_at': None,
            'rejection_reason': None,
            'attempts': [],               # [{submitted_at, action, actor, reason}]
        },

        # ── Site Verification ─────────────────────────────────────────────
        # FE captures 3 GPS-watermarked photos at site before creating tracker.
        # NOC confirms (FE present) or rejects (FE not at site) before assigning.
        # Rejected trackers require FE to resubmit new photos; rejection time is
        # excluded from NOC queue-wait KPI (queue wait = confirmed_at → assigned_at).
        'site_verification': {
            'status': 'pending',          # pending | confirmed | rejected
            'images': data.get('site_images', []),  # [{type, data, gps:{lat,lng,address}, captured_at}]
            'noc_reviewed_at': None,
            'noc_reviewed_by': None,
            'noc_reviewer_name': None,
            'rejection_reason': None,
            'rejection_count': 0,
            'last_submitted_at': now,
        },

        # ── Dedicated Stage Timestamps ────────────────────────────────────
        # These are set once (first occurrence) as the tracker progresses.
        # Having them as top-level indexed fields makes KPI aggregation fast
        # without scanning the events array on every analytics query.
        'stage_timestamps': {
            'tracker_created_at':                now,
            'site_verification_submitted_at':    now,
            'site_verification_confirmed_at':    None,
            'noc_assigned_at':                   None,
            'sim1_activation_started_at':        None,
            'sim1_activation_done_at':           None,
            'sim2_activation_started_at':        None,
            'sim2_activation_done_at':           None,
            'ztp_config_verified_at':            None,
            'ztp_started_at':                    None,
            'ztp_done_at':                       None,
            'ready_for_coordination_at':         None,
            'hso_submitted_at':                  None,
            'hso_approved_at':                   None,
            'installation_complete_at':          None,
        },

        # ── Event Log ─────────────────────────────────────────────────────
        # Append-only audit trail. Every action adds an event.
        'events': [make_event('tracker_created', session['user_id'], ROLE_FE,
                              'Installation tracker created at site')],

        # ── Status ────────────────────────────────────────────────────────
        'status': 'waiting_noc_assignment',
        'created_at': now,
        'updated_at': now,
        'completed_at': None,
    }

    result = mongo.db.trackers.insert_one(tracker)
    tracker['_id'] = result.inserted_id
    
    # Broadcast new tracker creation to all dashboards
    broadcast_dashboard_update(ROLE_NS, 'tracker_created', {
        'tracker_id': str(result.inserted_id),
        'sdwan_id': data['sdwan_id'],
        'customer': data['customer'],
        'status': 'waiting_noc_assignment'
    })
    
    # Also broadcast to FE hierarchy roles
    for role in [ROLE_FE, ROLE_FEG, ROLE_FS, ROLE_FSG]:
        broadcast_dashboard_update(role, 'tracker_created', {
            'tracker_id': str(result.inserted_id),
            'fe_id': session['user_id']
        })
    
    return jsonify({'success': True, 'tracker': serialize_doc(tracker)})


# ─── NOC: Assign Tracker ────────────────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/assign', methods=['POST'])
@login_required
def api_assign_tracker(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    # Guard: site verification must be confirmed before assignment
    sv = tracker.get('site_verification', {})
    if sv.get('status') != 'confirmed':
        return jsonify({'error': 'Site verification not confirmed. Confirm FE is at site before assigning.'}), 400

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'noc_assignee': session['user_id'],
                'status': 'noc_working',
                'stage_timestamps.noc_assigned_at': now,
                'updated_at': now
            },
            '$push': {
                'events': make_event('noc_assigned', session['user_id'], ROLE_NS,
                                     f"Assigned to {noc_name}",
                                     {'noc_name': noc_name}),
                'noc_history': {
                    'assignee_id': session['user_id'],
                    'assignee_name': noc_name,
                    'assigned_at': now,
                    'released_at': None
                }
            }
        }
    )
    
    # Broadcast assignment update via Socket.IO
    broadcast_tracker_update(tracker_id, 'tracker_assigned', {
        'noc_assignee': session['user_id'],
        'noc_name': noc_name,
        'status': 'noc_working'
    })
    
    # Broadcast to NOC dashboards
    broadcast_dashboard_update(ROLE_NS, 'tracker_assigned', {
        'tracker_id': tracker_id,
        'noc_assignee': session['user_id']
    })
    
    return jsonify({'success': True})


# ─── NOC: Confirm Site Verification ─────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/site-verify/confirm', methods=['POST'])
@login_required
def api_site_verify_confirm(tracker_id):
    if session.get('role') not in NOC_ROLES:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    sv = tracker.get('site_verification', {})
    if sv.get('status') != 'pending':
        return jsonify({'error': 'Site verification is not in pending state'}), 400

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'site_verification.status': 'confirmed',
                'site_verification.noc_reviewed_at': now,
                'site_verification.noc_reviewed_by': session['user_id'],
                'site_verification.noc_reviewer_name': noc_name,
                'stage_timestamps.site_verification_confirmed_at': now,
                'updated_at': now,
            },
            '$push': {
                'events': make_event('site_verification_confirmed', session['user_id'],
                                     session['role'],
                                     f"Site verification confirmed by {noc_name} — FE confirmed at site")
            }
        }
    )

    broadcast_tracker_update(tracker_id, 'site_verification_confirmed', {
        'site_verification_status': 'confirmed',
        'noc_reviewer_name': noc_name,
    })
    # Notify FE dashboard
    fe_id = tracker.get('fe', {}).get('id')
    if fe_id:
        broadcast_to_user(fe_id, 'site_verification_confirmed', {
            'tracker_id': tracker_id,
            'sdwan_id': tracker.get('sdwan_id'),
            'message': 'Your site verification was confirmed. Awaiting NOC assignment.'
        })

    return jsonify({'success': True})


# ─── NOC: Reject Site Verification ──────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/site-verify/reject', methods=['POST'])
@login_required
def api_site_verify_reject(tracker_id):
    if session.get('role') not in NOC_ROLES:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    sv = tracker.get('site_verification', {})
    if sv.get('status') != 'pending':
        return jsonify({'error': 'Site verification is not in pending state'}), 400

    data = request.json or {}
    reason = data.get('reason', '').strip()
    if not reason:
        return jsonify({'error': 'Rejection reason is required'}), 400

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'site_verification.status': 'rejected',
                'site_verification.noc_reviewed_at': now,
                'site_verification.noc_reviewed_by': session['user_id'],
                'site_verification.noc_reviewer_name': noc_name,
                'site_verification.rejection_reason': reason,
                'updated_at': now,
            },
            '$inc': {
                'site_verification.rejection_count': 1,
            },
            '$push': {
                'events': make_event('site_verification_rejected', session['user_id'],
                                     session['role'],
                                     f"Site verification rejected by {noc_name}: {reason}")
            }
        }
    )

    broadcast_tracker_update(tracker_id, 'site_verification_rejected', {
        'site_verification_status': 'rejected',
        'rejection_reason': reason,
    })
    # Notify FE with urgent push
    fe_id = tracker.get('fe', {}).get('id')
    if fe_id:
        broadcast_to_user(fe_id, 'site_verification_rejected', {
            'tracker_id': tracker_id,
            'sdwan_id': tracker.get('sdwan_id'),
            'rejection_reason': reason,
            'message': f"Site verification rejected: {reason}. Please retake site photos."
        })

    return jsonify({'success': True})


# ─── FE: Resubmit Site Verification ─────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/site-verify/resubmit', methods=['POST'])
@login_required
def api_site_verify_resubmit(tracker_id):
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Only Field Engineers can resubmit site verification'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    # Only the FE who created the tracker can resubmit
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Unauthorized — not the tracker owner'}), 403

    sv = tracker.get('site_verification', {})
    if sv.get('status') != 'rejected':
        return jsonify({'error': 'Site verification is not in rejected state'}), 400

    data = request.json or {}
    new_images = data.get('site_images', [])
    if len(new_images) < 3:
        return jsonify({'error': 'All 3 site photos are required'}), 400

    now = get_utc_now()

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'site_verification.status': 'pending',
                'site_verification.images': new_images,
                'site_verification.last_submitted_at': now,
                'site_verification.noc_reviewed_at': None,
                'site_verification.noc_reviewed_by': None,
                'site_verification.noc_reviewer_name': None,
                'site_verification.rejection_reason': None,
                'stage_timestamps.site_verification_submitted_at': now,
                'updated_at': now,
            },
            '$push': {
                'events': make_event('site_verification_resubmitted', session['user_id'],
                                     ROLE_FE,
                                     'FE resubmitted site verification photos')
            }
        }
    )

    broadcast_tracker_update(tracker_id, 'site_verification_resubmitted', {
        'site_verification_status': 'pending',
    })
    broadcast_dashboard_update(ROLE_NS, 'site_verification_resubmitted', {
        'tracker_id': tracker_id,
        'sdwan_id': tracker.get('sdwan_id'),
    })

    return jsonify({'success': True})


# ─── NOC: Request Reassignment ──────────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/request-reassignment', methods=['POST'])
@login_required
def api_request_reassignment(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    target_noc_id = data.get('target_noc_id', '').strip()
    reason = data.get('reason', '').strip()
    
    if not target_noc_id or not reason:
        return jsonify({'error': 'Target NOC user and reason are required'}), 400

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    # Only the currently assigned NOC can request reassignment
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Only the assigned NOC can request reassignment'}), 403

    # Cannot request to reassign to yourself
    if target_noc_id == session['user_id']:
        return jsonify({'error': 'Cannot request reassignment to yourself'}), 400

    # Get target NOC user details
    target_user = mongo.db.users.find_one({'_id': ObjectId(target_noc_id), 'role': ROLE_NS})
    if not target_user:
        return jsonify({'error': 'Target NOC user not found'}), 404

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    # Create reassignment request
    reassignment_request = {
        'from_noc_id': session['user_id'],
        'from_noc_name': noc_name,
        'to_noc_id': target_noc_id,
        'to_noc_name': target_user.get('name', target_user['username']),
        'reason': reason,
        'status': 'pending',  # pending, accepted, denied
        'requested_at': now,
        'responded_at': None,
        'response_reason': None
    }

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'reassignment_request': reassignment_request,
                'updated_at': now
            },
            '$push': {
                'events': make_event('reassignment_requested', session['user_id'], ROLE_NS,
                                     f"{noc_name} requested to transfer tracker to {reassignment_request['to_noc_name']}. Reason: {reason}",
                                     reassignment_request)
            }
        }
    )
    return jsonify({'success': True})


# ─── NOC: Accept Reassignment Request ───────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/accept-reassignment', methods=['POST'])
@login_required
def api_accept_reassignment(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    req = tracker.get('reassignment_request')
    if not req or req.get('status') != 'pending':
        return jsonify({'error': 'No pending reassignment request found'}), 404

    # Only the target NOC can accept
    if req.get('to_noc_id') != session['user_id']:
        return jsonify({'error': 'Only the target NOC can accept this request'}), 403

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    # Update request status
    req['status'] = 'accepted'
    req['responded_at'] = now

    # Perform the reassignment
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'noc_assignee': session['user_id'],
                'reassignment_request': req,
                'updated_at': now
            },
            '$push': {
                'events': make_event('reassignment_accepted', session['user_id'], ROLE_NS,
                                     f"{noc_name} accepted reassignment request from {req['from_noc_name']}",
                                     {'request': req}),
                'noc_history': {
                    'assignee_id': session['user_id'],
                    'assignee_name': noc_name,
                    'assigned_at': now,
                    'released_at': None,
                    'reassignment_reason': f"Accepted transfer from {req['from_noc_name']}: {req['reason']}",
                    'previous_assignee': req['from_noc_name']
                }
            }
        }
    )
    
    # Broadcast reassignment via Socket.IO
    broadcast_tracker_update(tracker_id, 'reassignment_accepted', {
        'new_assignee': session['user_id'],
        'new_assignee_name': noc_name,
        'previous_assignee': req['from_noc_id']
    })
    
    # Notify both NOC users
    broadcast_to_user(req['from_noc_id'], 'reassignment_accepted', {
        'tracker_id': tracker_id,
        'accepted_by': noc_name
    })
    broadcast_to_user(session['user_id'], 'reassignment_accepted', {
        'tracker_id': tracker_id,
        'message': 'You accepted the transfer request'
    })
    
    # Update NOC dashboards
    broadcast_dashboard_update(ROLE_NS, 'tracker_reassigned', {
        'tracker_id': tracker_id,
        'new_assignee': session['user_id']
    })
    
    return jsonify({'success': True})


# ─── NOC: Deny Reassignment Request ─────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/deny-reassignment', methods=['POST'])
@login_required
def api_deny_reassignment(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    denial_reason = data.get('reason', '').strip()
    if not denial_reason:
        return jsonify({'error': 'Denial reason is required'}), 400

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    req = tracker.get('reassignment_request')
    if not req or req.get('status') != 'pending':
        return jsonify({'error': 'No pending reassignment request found'}), 404

    # Only the target NOC can deny
    if req.get('to_noc_id') != session['user_id']:
        return jsonify({'error': 'Only the target NOC can deny this request'}), 403

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    # Update request status
    req['status'] = 'denied'
    req['responded_at'] = now
    req['response_reason'] = denial_reason

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'reassignment_request': req,
                'updated_at': now
            },
            '$push': {
                'events': make_event('reassignment_denied', session['user_id'], ROLE_NS,
                                     f"{noc_name} denied reassignment request from {req['from_noc_name']}. Reason: {denial_reason}",
                                     {'request': req, 'denial_reason': denial_reason})
            }
        }
    )
    return jsonify({'success': True})


# ─── NOC: Get Reassignment Requests ─────────────────────────────────────────
@app.route('/api/trackers/reassignment-requests', methods=['GET'])
@login_required
def api_get_reassignment_requests():
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    # Find all trackers with pending reassignment requests for current user
    trackers = list(mongo.db.trackers.find({
        'reassignment_request.to_noc_id': session['user_id'],
        'reassignment_request.status': 'pending'
    }))

    requests = []
    for t in trackers:
        req = t.get('reassignment_request', {})
        requests.append({
            'tracker_id': str(t['_id']),
            'sdwan_id': t.get('sdwan_id'),
            'customer': t.get('customer'),
            'from_noc_name': req.get('from_noc_name'),
            'reason': req.get('reason'),
            'requested_at': req.get('requested_at').isoformat() + 'Z' if req.get('requested_at') else None
        })

    return jsonify({'requests': requests})


# ─── NOC: Revoke Reassignment Request ───────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/revoke-reassignment', methods=['POST'])
@login_required
def api_revoke_reassignment(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    req = tracker.get('reassignment_request')
    if not req or req.get('status') != 'pending':
        return jsonify({'error': 'No pending reassignment request found'}), 404

    # Only the requesting NOC can revoke
    if req.get('from_noc_id') != session['user_id']:
        return jsonify({'error': 'Only the requesting NOC can revoke this request'}), 403

    now = get_utc_now()
    noc_name = session.get('noc_name', session['username'])

    # Update request status to revoked
    req['status'] = 'revoked'
    req['responded_at'] = now

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'reassignment_request': req,
                'updated_at': now
            },
            '$push': {
                'events': make_event('reassignment_revoked', session['user_id'], ROLE_NS,
                                     f"{noc_name} revoked transfer request to {req['to_noc_name']}",
                                     {'request': req})
            }
        }
    )
    return jsonify({'success': True})


# ─── NOC: SIM Activation ────────────────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/sim/<sim_key>/status', methods=['POST'])
@login_required
def api_update_sim_status(tracker_id, sim_key):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'This tracker is not assigned to you'}), 403

    data = request.json
    status         = data.get('status')
    failure_reason = data.get('failure_reason')
    remarks        = data.get('remarks', f'{sim_key.upper()} status → {status}')
    now            = get_utc_now()

    stage_map = {
        'activation_in_process':              f'{sim_key}_activation_in_process',
        'activation_complete_manual':         f'{sim_key}_activation_complete_manual',
        'activation_complete_preactivated':   f'{sim_key}_activation_complete_preactivated',
        'activation_failed':                  f'{sim_key}_activation_failed',
    }

    update_data = {f'sim.{sim_key}.status': status, 'updated_at': now}

    if status == 'activation_in_process':
        # Only stamp activation_started_at on the FIRST attempt
        if not tracker.get('sim', {}).get(sim_key, {}).get('activation_started_at'):
            update_data[f'sim.{sim_key}.activation_started_at'] = now
            update_data[f'stage_timestamps.{sim_key}_activation_started_at'] = now
        # Record the attempt starting
        attempt = {'attempt_no': len(tracker.get('sim', {}).get(sim_key, {}).get('attempts', [])) + 1,
                   'started_at': now, 'result': None, 'reason': None}
    elif status in ('activation_complete_manual', 'activation_complete_preactivated'):
        update_data[f'sim.{sim_key}.activation_done_at'] = now
        update_data[f'stage_timestamps.{sim_key}_activation_done_at'] = now
        if failure_reason:
            update_data[f'sim.{sim_key}.failure_reason'] = None  # Clear on success
    elif status == 'activation_failed':
        update_data[f'sim.{sim_key}.failure_reason'] = failure_reason

    event = make_event(stage_map.get(status, f'{sim_key}_status_update'),
                       session['user_id'], ROLE_NS, remarks,
                       {'sim_key': sim_key, 'status': status, 'failure_reason': failure_reason})

    # Append to the per-SIM attempts array for retry tracking
    attempt_entry = {
        'attempt_no': len(tracker.get('sim', {}).get(sim_key, {}).get('attempts', [])) + 1,
        'started_at': now if status == 'activation_in_process' else None,
        'result': None if status == 'activation_in_process' else status,
        'reason': failure_reason
    }

    push_ops = {'events': event}
    # Only push a new attempt record when we START an activation (not on every status change)
    if status == 'activation_in_process':
        push_ops[f'sim.{sim_key}.attempts'] = attempt_entry

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {'$set': update_data, '$push': push_ops}
    )
    
    # Broadcast SIM status update via Socket.IO
    broadcast_tracker_update(tracker_id, 'sim_status_updated', {
        'sim_key': sim_key,
        'status': status,
        'failure_reason': failure_reason
    })
    
    return jsonify({'success': True, 'message': f'{sim_key.upper()} status updated'})


# ─── NOC: ZTP Config Verification ───────────────────────────────────────────
# This is a NEW stage that was missing. NS verifies the ZTP configuration
# based on what FE submitted (firmware version, SDWAN ID etc.) before
# either NS or FE can initiate ZTP.
@app.route('/api/trackers/<tracker_id>/ztp/config', methods=['POST'])
@login_required
def api_verify_ztp_config(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    data = request.json
    result         = data.get('result')   # 'verified' or 'failed'
    failure_reason = data.get('failure_reason')
    remarks        = data.get('remarks', f'ZTP config {result}')
    now            = get_utc_now()

    if result not in ('verified', 'failed'):
        return jsonify({'error': 'result must be "verified" or "failed"'}), 400

    update_data = {
        'ztp.config_status': f'config_{result}',
        'stage_timestamps.ztp_config_verified_at': now,
        'updated_at': now
    }
    if result == 'failed':
        update_data['ztp.config_failure_reason'] = failure_reason

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': update_data,
            '$push': {'events': make_event(f'ztp_config_{result}', session['user_id'], ROLE_NS,
                                           remarks, {'failure_reason': failure_reason})}
        }
    )
    
    # Broadcast ZTP config verification via Socket.IO
    broadcast_tracker_update(tracker_id, 'ztp_config_updated', {
        'result': result,
        'failure_reason': failure_reason
    })
    
    return jsonify({'success': True, 'message': f'ZTP config marked as {result}'})


# ─── FE: Start ZTP (FE performs ZTP) ────────────────────────────────────────
# After NS verifies ZTP config, FE sees a "Start ZTP" button.
# FE clicks it, which initiates the ZTP process from the field device.
@app.route('/api/trackers/<tracker_id>/ztp/fe-start', methods=['POST'])
@login_required
def api_fe_start_ztp(tracker_id):
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403
    if tracker.get('ztp', {}).get('config_status') != 'config_verified':
        return jsonify({'error': 'ZTP config not yet verified by NOC'}), 400

    now = get_utc_now()
    attempt_no = len(tracker.get('ztp', {}).get('attempts', [])) + 1

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'ztp.status': 'initiated',
                'ztp.performed_by': 'FE',
                'ztp.initiated_at': now,
                'stage_timestamps.ztp_started_at': now,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_initiated_by_fe', session['user_id'], ROLE_FE,
                                     'FE initiated ZTP from device'),
                'ztp.attempts': {
                    'attempt_no': attempt_no,
                    'performed_by': 'FE',
                    'started_at': now,
                    'result': None,
                    'reason': None
                }
            }
        }
    )
    
    # Broadcast ZTP initiation via Socket.IO
    broadcast_tracker_update(tracker_id, 'ztp_initiated', {
        'performed_by': 'FE'
    })
    
    return jsonify({'success': True, 'message': 'ZTP initiated by FE'})


# ─── FE: ZTP Completed (FE reports ZTP succeeded) ───────────────────────────
@app.route('/api/trackers/<tracker_id>/ztp/fe-complete', methods=['POST'])
@login_required
def api_fe_complete_ztp(tracker_id):
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403

    now = get_utc_now()
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'ztp.status': 'fe_completed',   # Distinct from 'completed' so NS must verify
                'ztp.fe_completed_at': now,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_completed_by_fe', session['user_id'], ROLE_FE,
                                     'FE reports ZTP completed — awaiting NS verification')
            }
        }
    )
    
    # Broadcast ZTP completion via Socket.IO
    broadcast_tracker_update(tracker_id, 'ztp_completed_by_fe', {
        'status': 'fe_completed'
    })
    return jsonify({'success': True, 'message': 'ZTP reported complete by FE — awaiting NS verification'})


# ─── FE: Request NS to Perform ZTP ──────────────────────────────────────────
# If FE cannot perform ZTP (device issue, connectivity), FE clicks this button.
# Chat unlocks immediately so FE can explain the situation to NS.
@app.route('/api/trackers/<tracker_id>/ztp/request-noc', methods=['POST'])
@login_required
def api_request_noc_ztp(tracker_id):
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403

    now = get_utc_now()
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'ztp.fe_requested_ns': True,
                'status': 'fe_requested_ztp',   # Chat unlocks on this status
                'updated_at': now
            },
            '$push': {
                'events': make_event('fe_requested_noc_ztp', session['user_id'], ROLE_FE,
                                     'FE cannot perform ZTP — requested NS to do it. Chat unlocked.')
            }
        }
    )
    
    # Broadcast ZTP request to NOC via Socket.IO
    broadcast_tracker_update(tracker_id, 'ztp_requested_from_noc', {
        'status': 'ztp_pull_requested_from_noc'
    })
    
    return jsonify({'success': True, 'message': 'Request sent to NOC. Chat is now unlocked.'})


# ─── NOC: ZTP Status Update (NS performs or verifies ZTP) ───────────────────
@app.route('/api/trackers/<tracker_id>/ztp/status', methods=['POST'])
@login_required
def api_update_ztp_status(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    data           = request.json
    status         = data.get('status')   # 'initiated' | 'completed' | 'failed' | 'verified_fe_completion'
    failure_reason = data.get('failure_reason')
    remarks        = data.get('remarks', f'ZTP {status}')
    now            = get_utc_now()

    stage_map = {
        'initiated':              'ztp_initiated_by_noc',
        'completed':              'ztp_completed_by_noc',
        'failed':                 'ztp_failed',
        'verified_fe_completion': 'ztp_fe_completion_verified',
    }

    update_data = {'ztp.status': status, 'updated_at': now}

    if status == 'initiated':
        update_data['ztp.performed_by'] = 'NS'
        update_data['ztp.initiated_at'] = now
        update_data['stage_timestamps.ztp_started_at'] = now
        attempt_no = len(tracker.get('ztp', {}).get('attempts', [])) + 1
        push_attempt = {
            'attempt_no': attempt_no,
            'performed_by': 'NS',
            'started_at': now,
            'result': None,
            'reason': None
        }
    elif status in ('completed', 'verified_fe_completion'):
        update_data['ztp.status'] = 'completed'  # Normalise both to 'completed'
        update_data['ztp.completed_at'] = now
        update_data['stage_timestamps.ztp_done_at'] = now
        push_attempt = None
    elif status == 'failed':
        update_data['ztp.failure_reason'] = failure_reason
        push_attempt = None

    event_stage = stage_map.get(status, 'ztp_status_update')
    push_ops = {'events': make_event(event_stage, session['user_id'], ROLE_NS, remarks,
                                     {'status': status, 'failure_reason': failure_reason})}
    if status == 'initiated' and push_attempt:
        push_ops['ztp.attempts'] = push_attempt

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {'$set': update_data, '$push': push_ops}
    )
    
    # Broadcast ZTP status update via Socket.IO
    broadcast_tracker_update(tracker_id, 'ztp_status_updated', {
        'status': status,
        'failure_reason': failure_reason
    })
    
    return jsonify({'success': True, 'message': f'ZTP status updated to {status}'})


# ─── NEW ZTP WORKFLOW ENDPOINTS (2-Phase: Config Verification → Pull Execution) ───

# Phase 1a: FE submits ZTP configuration for NOC verification
@app.route('/api/ztp/config/submit', methods=['POST'])
@login_required
def api_ztp_config_submit():
    """FE submits ZTP device configuration for NOC to verify."""
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    tracker_id = data.get('tracker_id')

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403

    now = get_utc_now()
    config_data = {
        'device_model': data.get('device_model', ''),
        'firmware_version': data.get('firmware_version', ''),
        'serial_number': data.get('serial_number', ''),
        'mac_address': data.get('mac_address', ''),
        'submitted_by': session['username'],
        'submitted_at': now,
    }

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'ztp.config': config_data,
                'ztp.config_status': 'pending',
                'status': STATUS_ZTP_PULL_PENDING if tracker.get('status') == STATUS_ZTP_CONFIG_UNVERIFIED else tracker.get('status'),
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_config_submitted', session['user_id'], ROLE_FE,
                                     'ZTP device configuration submitted for NOC verification',
                                     {'config': config_data})
            }
        }
    )

    return jsonify({'success': True, 'message': 'Configuration submitted for NOC verification'})


# Phase 1b: NOC verifies ZTP Configuration
@app.route('/api/ztp/config/verify', methods=['POST'])
@login_required
def api_ztp_config_verify():
    """NOC verifies or rejects ZTP configuration"""
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    tracker_id = data.get('tracker_id')
    # Accept both 'result' (canonical) and 'verified' (from ZTP component JS boolean)
    result = data.get('result')
    if result is None:
        verified_flag = data.get('verified')
        if verified_flag is True:
            result = 'verified'
        elif verified_flag is False:
            result = 'unverified'
    notes = data.get('notes', '')

    if result not in ('verified', 'unverified'):
        return jsonify({'error': 'result must be "verified" or "unverified"'}), 400

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    now = get_utc_now()

    if result == 'verified':
        new_status = STATUS_ZTP_PULL_PENDING
        event_msg = 'ZTP Configuration Verified by NOC'
    else:
        new_status = STATUS_ZTP_CONFIG_UNVERIFIED
        event_msg = f'ZTP Configuration Marked as Unverified: {notes}'

    verification_entry = {
        'verified_by': session['username'],
        'result': result,
        'timestamp': now,
        'notes': notes
    }

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'status': new_status,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_config_verification', session['user_id'], ROLE_NS, event_msg),
                'ztp.config_verification_history': verification_entry
            }
        }
    )

    return jsonify({'success': True, 'message': f'ZTP config marked as {result}'})


# Phase 2: ZTP Pull Execution - FE Actions
@app.route('/api/ztp/pull/action', methods=['POST'])
@login_required
def api_ztp_pull_action():
    """FE marks ZTP pull as done or requests NOC to perform it"""
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    tracker_id = data.get('tracker_id')
    action = data.get('action')  # 'done_by_fe' or 'request_noc'
    notes = data.get('notes', '')

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403

    now = get_utc_now()

    if action == 'done_by_fe':
        new_status = STATUS_ZTP_PULL_DONE_FE
        event_msg = 'FE Completed ZTP Pull - Awaiting NOC Verification'
        pull_entry = {
            'action': 'done_by_fe',
            'performed_by': session['username'],
            'timestamp': now,
            'notes': notes
        }
    elif action == 'request_noc':
        new_status = STATUS_ZTP_PULL_REQ_NOC
        event_msg = 'FE Requested NOC to Perform ZTP Pull'
        pull_entry = {
            'action': 'requested_from_noc',
            'performed_by': session['username'],
            'timestamp': now,
            'notes': notes
        }
    else:
        return jsonify({'error': 'Invalid action'}), 400

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'status': new_status,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_pull_action', session['user_id'], ROLE_FE, event_msg),
                'ztp.pull_history': pull_entry
            }
        }
    )

    return jsonify({'success': True, 'message': event_msg})


# Phase 2: ZTP Pull Execution - NOC Verifies FE's Work
@app.route('/api/ztp/pull/verify', methods=['POST'])
@login_required
def api_ztp_pull_verify():
    """NOC verifies or rejects FE's ZTP pull work"""
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    tracker_id = data.get('tracker_id')
    # Accept both 'result' (canonical) and 'verified' (from ZTP component JS boolean)
    result = data.get('result')
    if result is None:
        verified_flag = data.get('verified')
        if verified_flag is True:
            result = 'verified'
        elif verified_flag is False:
            result = 'unverified'
    notes = data.get('notes', '')

    if result not in ('verified', 'unverified'):
        return jsonify({'error': 'result must be "verified" or "unverified"'}), 400

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    now = get_utc_now()

    pull_entry = {
        'action': result,
        'performed_by': session['username'],
        'timestamp': now,
        'notes': notes
    }

    if result == 'verified':
        event_msg = "NOC Verified FE's ZTP Pull - ZTP Complete"
        # Mirror what api_update_ztp_status does for status='verified_fe_completion':
        # Set ztp.status='completed', write all stage timestamps, and advance the
        # tracker to ready_for_coordination so the downstream chain (Quick Actions,
        # HSO submit button, chat unlock) all work correctly.
        set_fields = {
            'status': STATUS_READY_COORD,
            'ztp.status': 'completed',
            'ztp.performed_by': 'FE',
            'ztp.completed_at': now,
            'stage_timestamps.ztp_done_at': now,
            'stage_timestamps.ready_for_coordination_at': now,
            'updated_at': now
        }
    else:
        event_msg = f'NOC Marked ZTP Pull as Unverified: {notes}'
        set_fields = {
            'status': STATUS_ZTP_PULL_UNVERIFIED,
            'updated_at': now
        }

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': set_fields,
            '$push': {
                'events': make_event('ztp_pull_verification', session['user_id'], ROLE_NS, event_msg),
                'ztp.pull_history': pull_entry
            }
        }
    )

    return jsonify({'success': True, 'message': event_msg})


# Phase 2: ZTP Pull Execution - NOC Performs ZTP
@app.route('/api/ztp/pull/perform-by-noc', methods=['POST'])
@login_required
def api_ztp_pull_perform_by_noc():
    """NOC performs ZTP pull when requested by FE"""
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json
    tracker_id = data.get('tracker_id')
    notes = data.get('notes', '')

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    now = get_utc_now()
    event_msg = 'ZTP Pull Completed by NOC - ZTP Complete'

    pull_entry = {
        'action': 'done_by_noc',
        'performed_by': session['username'],
        'timestamp': now,
        'notes': notes
    }

    noc_name = session.get('noc_name', session['username'])

    # Transition the tracker to ready_for_coordination so that:
    #   - The Installation Status tile shows "Ready — Chat Open" (not "FE Needs Help")
    #   - is_chat_unlocked() returns True, enabling the chat
    #   - renderQuickActions() shows the HSO Approve/Reject section
    #   - renderHsoSection() shows the FE's Submit HSO button
    # Also write all ZTP completion fields that the rest of the code expects:
    #   - ztp.status = 'completed'  (renderQuickActions checks ztpDone = ztp.status === 'completed')
    #   - ztp.performed_by = 'NS'   (KPI: ZTP done by FE vs NS)
    #   - ztp.completed_at          (stage duration KPI)
    #   - stage_timestamps.ztp_done_at  (fast KPI aggregation field)
    #   - stage_timestamps.ready_for_coordination_at  (NS processing time KPI)
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'status': STATUS_READY_COORD,
                'ztp.status': 'completed',
                'ztp.performed_by': 'NS',
                'ztp.completed_at': now,
                'stage_timestamps.ztp_done_at': now,
                'stage_timestamps.ready_for_coordination_at': now,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ztp_completed_by_noc', session['user_id'], ROLE_NS, event_msg,
                                     {'performed_by': 'NS', 'notes': notes}),
                'ztp.pull_history': pull_entry,
                # Also record in the canonical ztp.attempts array for KPI tracking
                'ztp.attempts': {
                    'attempt_no': len(tracker.get('ztp', {}).get('attempts', [])) + 1,
                    'performed_by': 'NS',
                    'started_at': now,
                    'result': 'completed',
                    'reason': notes or None
                }
            }
        }
    )
    
    # Broadcast ready for coordination via Socket.IO
    broadcast_tracker_update(tracker_id, 'ready_for_coordination', {
        'status': 'ready_for_coordination'
    })

    return jsonify({'success': True, 'message': 'ZTP pull completed by NOC. Ready for coordination.'})


# ─── NOC: Mark Ready for Coordination ───────────────────────────────────────
# Called after ZTP is fully completed (by FE verified by NS, or by NS directly).
# This unlocks the chat channel for FE-NS coordination.
@app.route('/api/trackers/<tracker_id>/ready-for-coordination', methods=['POST'])
@login_required
def api_ready_for_coordination(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    now = get_utc_now()
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'status': 'ready_for_coordination',
                'stage_timestamps.ready_for_coordination_at': now,
                'updated_at': now
            },
            '$push': {
                'events': make_event('ready_for_coordination', session['user_id'], ROLE_NS,
                                     'SIM activation & ZTP complete. Chat unlocked. Ready for FE coordination.')
            }
        }
    )
    return jsonify({'success': True, 'message': 'Marked as ready for coordination. Chat is now unlocked.'})


# ─── FE: Submit HSO ─────────────────────────────────────────────────────────
# FE clicks this button to formally notify NS that the hand-over sign-off
# document has been submitted. This creates a timestamped event and moves
# the tracker to 'hso_submitted' status so NS can see it.
@app.route('/api/trackers/<tracker_id>/hso/submit', methods=['POST'])
@login_required
def api_submit_hso(tracker_id):
    if session.get('role') != ROLE_FE:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403
    # Allow HSO submission from every status where the FE-NS coordination phase has begun.
    if tracker.get('status') not in HSO_SUBMITTABLE_STATUSES:
        return jsonify({'error': 'Tracker is not in the correct state to submit HSO'}), 400

    now = get_utc_now()
    attempt_no = len(tracker.get('hso', {}).get('attempts', [])) + 1
    # First submission sets submitted_at; resubmissions add to the attempts log
    set_ops = {
        'hso.status': 'submitted',
        'hso.submitted_at': now,   # Tracks the LATEST submission timestamp
        'status': 'hso_submitted',
        'updated_at': now
    }
    # Only stamp the KPI field on the very first submission
    if attempt_no == 1:
        set_ops['stage_timestamps.hso_submitted_at'] = now

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': set_ops,
            '$push': {
                'events': make_event('hso_submitted', session['user_id'], ROLE_FE,
                                     f'HSO submitted by FE (attempt #{attempt_no})'),
                'hso.attempts': {
                    'attempt_no': attempt_no,
                    'submitted_at': now,
                    'action': 'submitted',
                    'actor_id': session['user_id'],
                    'actor_role': ROLE_FE,
                    'reason': None
                }
            }
        }
    )
    
    # Broadcast HSO submission via Socket.IO
    broadcast_tracker_update(tracker_id, 'hso_submitted', {
        'status': 'hso_submitted',
        'attempt_no': attempt_no
    })
    
    return jsonify({'success': True, 'message': f'HSO submitted (attempt #{attempt_no})'})


# ─── NOC: Approve HSO ───────────────────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/hso/approve', methods=['POST'])
@login_required
def api_approve_hso(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403
    if tracker.get('hso', {}).get('status') != 'submitted':
        return jsonify({'error': 'No pending HSO submission to approve'}), 400

    data    = request.json or {}
    remarks = data.get('remarks', 'HSO approved by NS')
    now     = get_utc_now()

    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'hso.status': 'approved',
                'hso.approved_at': now,
                'status': 'installation_complete',
                'stage_timestamps.hso_approved_at': now,
                'stage_timestamps.installation_complete_at': now,
                'completed_at': now,
                'updated_at': now
            },
            '$push': {
                'events': {
                    '$each': [
                        make_event('hso_approved', session['user_id'], ROLE_NS, remarks),
                        make_event('installation_complete', 'system', 'system',
                                   'Installation completed successfully'),
                    ]
                }
            }
        }
    )
    
    # Broadcast HSO approval and installation completion via Socket.IO
    broadcast_tracker_update(tracker_id, 'installation_complete', {
        'status': 'installation_complete',
        'hso_status': 'approved'
    })
    
    return jsonify({'success': True, 'message': 'HSO approved. Installation complete!'})


# ─── NOC: Reject HSO ────────────────────────────────────────────────────────
# NS rejects the HSO with a reason. FE will see the rejection reason and
# a "Re-submit HSO" button. The tracker reverts to 'hso_rejected' status.
@app.route('/api/trackers/<tracker_id>/hso/reject', methods=['POST'])
@login_required
def api_reject_hso(tracker_id):
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403

    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403
    if tracker.get('hso', {}).get('status') != 'submitted':
        return jsonify({'error': 'No pending HSO submission to reject'}), 400

    data   = request.json or {}
    reason = data.get('reason')
    if not reason:
        return jsonify({'error': 'Rejection reason is required'}), 400

    now = get_utc_now()
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'hso.status': 'rejected',
                'hso.rejected_at': now,
                'hso.rejection_reason': reason,
                'status': 'hso_rejected',   # FE sees this and can Re-submit
                'updated_at': now
            },
            '$push': {
                'events': make_event('hso_rejected', session['user_id'], ROLE_NS,
                                     f'HSO rejected: {reason}', {'reason': reason})
            }
        }
    )
    
    # Broadcast HSO rejection via Socket.IO
    broadcast_tracker_update(tracker_id, 'hso_rejected', {
        'status': 'hso_rejected',
        'reason': reason
    })
    
    return jsonify({'success': True, 'message': 'HSO rejected. FE will be notified to re-submit.'})


# ─── NOC: Legacy "incomplete" endpoint ──────────────────────────────────────
# Kept for backwards compatibility but internally maps to the reject flow.
@app.route('/api/trackers/<tracker_id>/hso/incomplete', methods=['POST'])
@login_required
def api_hso_incomplete(tracker_id):
    """Legacy endpoint — routes to reject internally."""
    if session.get('role') != ROLE_NS:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json or {}
    data['reason'] = data.get('reason') or data.get('remarks', 'HSO incomplete')
    # Re-use reject handler logic inline
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    if tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    reason = data['reason']
    now = get_utc_now()
    mongo.db.trackers.update_one(
        {'_id': ObjectId(tracker_id)},
        {
            '$set': {
                'hso.status': 'rejected',
                'hso.rejected_at': now,
                'hso.rejection_reason': reason,
                'status': 'hso_rejected',
                'updated_at': now
            },
            '$push': {
                'events': make_event('hso_rejected', session['user_id'], ROLE_NS,
                                     f'HSO incomplete/rejected: {reason}', {'reason': reason})
            }
        }
    )
    return jsonify({'success': True, 'message': 'HSO marked as incomplete (rejected)'})


# ─── Predefined Reasons ─────────────────────────────────────────────────────
@app.route('/api/config/reasons/<category>')
@login_required
def api_get_reasons(category):
    doc = mongo.db.predefined_reasons.find_one({'category': category})
    return jsonify({'reasons': doc.get('reasons', []) if doc else []})


# ─── Chat APIs ──────────────────────────────────────────────────────────────
@app.route('/api/trackers/<tracker_id>/chat/messages', methods=['GET'])
@login_required
def api_get_chat_messages(tracker_id):
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    user_role = session.get('role')
    # NOC operators can only see chat for trackers assigned to them
    if user_role == ROLE_NS and tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    can_interact = True
    if user_role == ROLE_FE:
        can_interact = (tracker.get('fe', {}).get('id') == session['user_id'])
    elif user_role in (ROLE_FEG, ROLE_FS, ROLE_FSG):
        can_interact = False  # Hierarchy supervisors are read-only

    messages = list(mongo.db.chat_messages.find({'tracker_id': tracker_id}).sort('timestamp', 1))
    return jsonify({
        'messages': serialize_doc(messages),
        'chat_unlocked': is_chat_unlocked(tracker),
        'can_interact': can_interact
    })


@app.route('/api/trackers/<tracker_id>/chat/send', methods=['POST'])
@login_required
def api_send_chat_message(tracker_id):
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    user_role = session.get('role')
    if user_role == ROLE_FE and tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403
    if user_role == ROLE_NS and tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403

    if not is_chat_unlocked(tracker):
        return jsonify({'error': 'Chat is locked. Complete SIM activation and ZTP first.'}), 403

    data = request.json
    message_text = data.get('message', '').strip()
    message_type = data.get('type', 'text')
    file_url     = data.get('file_url')

    if not message_text and not file_url:
        return jsonify({'error': 'Message or file required'}), 400

    now = get_utc_now()
    sender_name = session.get('noc_name') or session.get('name') or session.get('username', 'User')
    message = {
        'tracker_id': tracker_id,
        'sender_id': session['user_id'],
        'sender_role': user_role,
        'sender_name': sender_name,
        'message': message_text,
        'type': message_type,
        'file_url': file_url,
        'timestamp': now,
        'read': False
    }

    result = mongo.db.chat_messages.insert_one(message)
    message['_id'] = result.inserted_id
    mongo.db.trackers.update_one({'_id': ObjectId(tracker_id)}, {'$set': {'updated_at': now}})
    
    # Broadcast message via Socket.IO for real-time delivery
    broadcast_chat_message(tracker_id, message)
    
    return jsonify({'success': True, 'message': serialize_doc(message)})


@app.route('/api/trackers/<tracker_id>/chat/upload', methods=['POST'])
@login_required
def api_upload_chat_file(tracker_id):
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404

    user_role = session.get('role')
    if user_role == ROLE_FE and tracker.get('fe', {}).get('id') != session['user_id']:
        return jsonify({'error': 'Not your tracker'}), 403
    if user_role == ROLE_NS and tracker.get('noc_assignee') != session['user_id']:
        return jsonify({'error': 'Not assigned to you'}), 403
    if not is_chat_unlocked(tracker):
        return jsonify({'error': 'Chat is locked'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file      = request.files['file']
    file_type = request.form.get('type', 'image')

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    import base64
    from io import BytesIO

    file_data = file.read()

    if file_type == 'image':
        # Sanity check — reject HTML blobs masquerading as images
        if file_data[:5] in (b'<!DOC', b'<html') or b'<!DOCTYPE' in file_data[:100]:
            return jsonify({'error': 'Invalid file type. Please upload an image file.'}), 400
        try:
            from PIL import Image
            img = Image.open(BytesIO(file_data))
            if img.mode in ('RGBA', 'LA', 'P'):
                bg = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                bg.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = bg
            max_size = 1024
            if max(img.size) > max_size:
                ratio = max_size / max(img.size)
                img = img.resize(tuple(int(d * ratio) for d in img.size), Image.Resampling.LANCZOS)
            out = BytesIO()
            img.save(out, format='JPEG', quality=85, optimize=True)
            file_data = out.getvalue()
            mime_type = 'image/jpeg'
        except ImportError:
            mime_type = file.content_type or 'image/jpeg'
        except Exception as e:
            return jsonify({'error': f'Failed to process image: {str(e)}'}), 400
    elif file_type == 'audio':
        mime_type = file.content_type or 'audio/webm'
    else:
        mime_type = file.content_type or 'application/octet-stream'

    b64 = base64.b64encode(file_data).decode('utf-8')
    file_url = f"data:{mime_type};base64,{b64}"

    if len(file_url) > 10 * 1024 * 1024:
        return jsonify({'error': 'File too large after encoding. Use a smaller file.'}), 400

    return jsonify({'success': True, 'file_url': file_url, 'type': file_type})


@app.route('/api/trackers/<tracker_id>/chat/mark-read', methods=['POST'])
@login_required
def api_mark_messages_read(tracker_id):
    tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
    if not tracker:
        return jsonify({'error': 'Tracker not found'}), 404
    mongo.db.chat_messages.update_many(
        {'tracker_id': tracker_id, 'sender_role': {'$ne': session.get('role')}, 'read': False},
        {'$set': {'read': True}}
    )
    return jsonify({'success': True})


# ─── Analytics / KPI APIs ───────────────────────────────────────────────────
def get_date_range(range_type, custom_from=None, custom_to=None):
    now = get_utc_now()
    if range_type == 'today':
        return now.replace(hour=0, minute=0, second=0, microsecond=0), now
    elif range_type == 'yesterday':
        y = now - timedelta(days=1)
        return y.replace(hour=0, minute=0, second=0, microsecond=0), \
               y.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif range_type == 'last7days':
        return (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0), now
    elif range_type == 'lastmonth':
        return (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0), now
    elif range_type == 'custom' and custom_from and custom_to:
        return (datetime.fromisoformat(custom_from.replace('Z', '+00:00')).replace(tzinfo=None),
                datetime.fromisoformat(custom_to.replace('Z', '+00:00')).replace(tzinfo=None))
    return now.replace(hour=0, minute=0, second=0, microsecond=0), now


def duration_minutes(t_start, t_end):
    if t_start and t_end:
        return max(0, (t_end - t_start).total_seconds() / 60)
    return None


def calculate_stage_times(tracker):
    """Compute KPI durations from the dedicated stage_timestamps sub-document.
    Falls back to scanning events only for fields not yet stamped in older records."""
    ts = tracker.get('stage_timestamps', {})

    def _ts(key):
        """Return the datetime for a timestamp key, handling both datetime and string formats."""
        val = ts.get(key)
        if isinstance(val, str):
            return datetime.fromisoformat(val.rstrip('Z'))
        return val

    created_at  = tracker.get('created_at') or _ts('tracker_created_at')
    assigned_at = _ts('noc_assigned_at')
    sim1_start  = _ts('sim1_activation_started_at')
    sim1_done   = _ts('sim1_activation_done_at')
    sim2_start  = _ts('sim2_activation_started_at')
    sim2_done   = _ts('sim2_activation_done_at')
    ztp_config  = _ts('ztp_config_verified_at')
    ztp_start   = _ts('ztp_started_at')
    ztp_done    = _ts('ztp_done_at')
    coord_at    = _ts('ready_for_coordination_at')
    hso_sub_at  = _ts('hso_submitted_at')
    hso_done_at = _ts('hso_approved_at')
    complete_at = tracker.get('completed_at') or _ts('installation_complete_at')

    stage_times = {
        'queue_wait_minutes':         duration_minutes(created_at, assigned_at),
        'sim1_activation_minutes':    duration_minutes(sim1_start, sim1_done),
        'sim2_activation_minutes':    duration_minutes(sim2_start, sim2_done),
        'ztp_config_minutes':         duration_minutes(assigned_at, ztp_config),
        'ztp_execution_minutes':      duration_minutes(ztp_start, ztp_done),
        'ns_processing_minutes':      duration_minutes(assigned_at, coord_at),
        'hso_review_minutes':         duration_minutes(hso_sub_at, hso_done_at),
        'total_minutes':              duration_minutes(created_at, complete_at),
    }
    return stage_times


def _analytics_allowed():
    return session.get('role') in {ROLE_ANALYTICS, ROLE_NSG, ROLE_FSG}


@app.route('/api/NOC_SUPPORT_GROUP/kpi')
@app.route('/api/analytics/kpi')
@login_required
def api_analytics_kpi():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403

    range_type  = request.args.get('range', 'today')
    custom_from = request.args.get('from')
    custom_to   = request.args.get('to')
    start, end  = get_date_range(range_type, custom_from, custom_to)

    started    = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}))
    completed  = list(mongo.db.trackers.find({'completed_at': {'$gte': start, '$lte': end}}))
    in_progress = mongo.db.trackers.count_documents({'created_at': {'$lte': end}, 'completed_at': None, 'noc_assignee': {'$ne': None}})
    unassigned  = mongo.db.trackers.count_documents({'completed_at': None, 'noc_assignee': None})

    # Average completion time and full stage breakdown over completed trackers
    completion_hours = []
    stage_avgs = {
        'queue_wait_minutes': [],
        'sim1_activation_minutes': [],
        'sim2_activation_minutes': [],
        'ztp_config_minutes': [],
        'ztp_execution_minutes': [],
        'ns_processing_minutes': [],
        'hso_review_minutes': [],
    }

    for t in completed:
        if t.get('created_at') and t.get('completed_at'):
            completion_hours.append((t['completed_at'] - t['created_at']).total_seconds() / 3600)
        st = calculate_stage_times(t)
        for k in stage_avgs:
            if st.get(k) is not None:
                stage_avgs[k].append(st[k])

    def _avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else None

    # SIM and ZTP quality metrics
    date_filter = {'created_at': {'$gte': start, '$lte': end}}
    sim1_fails = mongo.db.trackers.count_documents({**date_filter, 'sim.sim1.failure_reason': {'$ne': None}})
    sim2_fails = mongo.db.trackers.count_documents({**date_filter, 'sim.sim2.failure_reason': {'$ne': None}})
    ztp_fails  = mongo.db.trackers.count_documents({**date_filter, 'ztp.failure_reason': {'$ne': None}})
    ztp_by_fe  = mongo.db.trackers.count_documents({**date_filter, 'ztp.performed_by': 'FE'})
    ztp_by_ns  = mongo.db.trackers.count_documents({**date_filter, 'ztp.performed_by': 'NS'})
    hso_multi  = mongo.db.trackers.count_documents({**date_filter, 'hso.attempts.1': {'$exists': True}})

    # Completion rate
    completion_rate = round(len(completed) / len(started) * 100, 1) if started else 0

    avg_h = _avg(completion_hours)
    return jsonify({
        'total_started':          len(started),
        'total_completed':        len(completed),
        'completion_rate':        completion_rate,
        'avg_completion':         f"{avg_h}h" if avg_h else "-",
        'avg_completion_hours':   avg_h,
        'in_progress':            in_progress,
        'unassigned':             unassigned,
        'sim1_failure_count':     sim1_fails,
        'sim2_failure_count':     sim2_fails,
        'sim_failure_count':      sim1_fails + sim2_fails,  # backward compat
        'ztp_failure_count':      ztp_fails,
        'ztp_by_fe':              ztp_by_fe,
        'ztp_by_ns':              ztp_by_ns,
        'hso_multi_attempt':      hso_multi,
        'avg_queue_wait_min':     _avg(stage_avgs['queue_wait_minutes']),
        'avg_sim1_act_min':       _avg(stage_avgs['sim1_activation_minutes']),
        'avg_sim2_act_min':       _avg(stage_avgs['sim2_activation_minutes']),
        'avg_ztp_config_min':     _avg(stage_avgs['ztp_config_minutes']),
        'avg_ztp_exec_min':       _avg(stage_avgs['ztp_execution_minutes']),
        'avg_ns_processing_min':  _avg(stage_avgs['ns_processing_minutes']),
        'avg_hso_review_min':     _avg(stage_avgs['hso_review_minutes']),
    })


@app.route('/api/NOC_SUPPORT_GROUP/fe/overview')
@app.route('/api/analytics/fe/overview')
@login_required
def api_analytics_fe_overview():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type  = request.args.get('range', 'today')
    start, end  = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    trackers    = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}))

    day_data = {}
    fe_accounts = set()
    for t in trackers:
        fe_user = t.get('fe', {}).get('username', 'Unknown')
        fe_accounts.add(fe_user)
        day_key = t['created_at'].strftime('%Y-%m-%d')
        day_data.setdefault(day_key, {}).setdefault(fe_user, {'started': 0, 'completed': 0})
        day_data[day_key][fe_user]['started'] += 1
        if t.get('completed_at'):
            day_data[day_key][fe_user]['completed'] += 1

    sorted_days = sorted(day_data.keys())
    fe_accounts = sorted(fe_accounts)
    colors = ['#3B82F6','#10B981','#F59E0B','#EF4444','#8B5CF6','#EC4899']
    datasets = []
    for i, fe in enumerate(fe_accounts):
        datasets.append({'label': f'{fe} - Completed',
                         'data': [day_data[d].get(fe, {'completed':0})['completed'] for d in sorted_days],
                         'backgroundColor': colors[i % len(colors)], 'stack': fe})
        datasets.append({'label': f'{fe} - In Progress',
                         'data': [day_data[d].get(fe, {'started':0,'completed':0})['started'] -
                                  day_data[d].get(fe, {'started':0,'completed':0})['completed'] for d in sorted_days],
                         'backgroundColor': '#' + hex(int(colors[i % len(colors)][1:], 16) + 0x333333)[2:].zfill(6),
                         'stack': fe})
    return jsonify({'labels': sorted_days, 'datasets': datasets})


@app.route('/api/NOC_SUPPORT_GROUP/noc/overview')
@app.route('/api/analytics/noc/overview')
@login_required
def api_analytics_noc_overview():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    trackers   = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}))

    day_data = {}
    for t in trackers:
        day_key = t['created_at'].strftime('%Y-%m-%d')
        day_data.setdefault(day_key, {'started': 0, 'completed': 0})
        day_data[day_key]['started'] += 1
        if t.get('completed_at'):
            day_data[day_key]['completed'] += 1

    sorted_days = sorted(day_data.keys())
    return jsonify({
        'labels': sorted_days,
        'datasets': [
            {'label': 'Completed',  'data': [day_data[d]['completed'] for d in sorted_days],
             'backgroundColor': '#10B981', 'stack': 'total'},
            {'label': 'In Progress','data': [day_data[d]['started'] - day_data[d]['completed'] for d in sorted_days],
             'backgroundColor': '#F59E0B', 'stack': 'total'},
        ]
    })


@app.route('/api/NOC_SUPPORT_GROUP/trend')
@app.route('/api/analytics/trend')
@login_required
def api_analytics_trend():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    trackers   = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}))

    day_data = {}
    for t in trackers:
        day_key = t['created_at'].strftime('%Y-%m-%d')
        day_data.setdefault(day_key, {'started': 0, 'completed': 0, 'times': []})
        day_data[day_key]['started'] += 1
        if t.get('completed_at'):
            day_data[day_key]['completed'] += 1
            day_data[day_key]['times'].append((t['completed_at'] - t['created_at']).total_seconds() / 3600)

    sorted_days = sorted(day_data.keys())
    return jsonify({
        'labels': sorted_days,
        'datasets': [
            {'label': 'Started',   'data': [day_data[d]['started'] for d in sorted_days],
             'type': 'bar', 'backgroundColor': '#3B82F6', 'yAxisID': 'y'},
            {'label': 'Completed', 'data': [day_data[d]['completed'] for d in sorted_days],
             'type': 'bar', 'backgroundColor': '#10B981', 'yAxisID': 'y'},
            {'label': 'Avg Completion (h)',
             'data': [round(sum(day_data[d]['times'])/len(day_data[d]['times']), 2)
                      if day_data[d]['times'] else 0 for d in sorted_days],
             'type': 'line', 'borderColor': '#F59E0B', 'backgroundColor': 'transparent',
             'yAxisID': 'y1', 'tension': 0.4},
        ]
    })


@app.route('/api/analytics/stage-durations')
@login_required
def api_analytics_stage_durations():
    """Average time spent in each stage (for pipeline / funnel charts)."""
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    completed = list(mongo.db.trackers.find({'completed_at': {'$gte': start, '$lte': end}}))

    stage_totals = {
        'Queue Wait': [], 'SIM1 Activation': [], 'SIM2 Activation': [],
        'ZTP Config': [], 'ZTP Execution': [], 'NS Processing': [], 'HSO Review': []
    }
    key_map = {
        'Queue Wait': 'queue_wait_minutes', 'SIM1 Activation': 'sim1_activation_minutes',
        'SIM2 Activation': 'sim2_activation_minutes', 'ZTP Config': 'ztp_config_minutes',
        'ZTP Execution': 'ztp_execution_minutes', 'NS Processing': 'ns_processing_minutes',
        'HSO Review': 'hso_review_minutes'
    }
    for t in completed:
        st = calculate_stage_times(t)
        for label, key in key_map.items():
            if st.get(key) is not None:
                stage_totals[label].append(st[key])

    def _avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else 0

    labels = list(stage_totals.keys())
    values = [_avg(stage_totals[l]) for l in labels]
    colors = ['#6366F1', '#3B82F6', '#0EA5E9', '#14B8A6', '#22C55E', '#F59E0B', '#EF4444']

    return jsonify({
        'labels': labels,
        'datasets': [{
            'label': 'Avg Duration (min)',
            'data': values,
            'backgroundColor': colors,
        }]
    })


@app.route('/api/analytics/status-distribution')
@login_required
def api_analytics_status_distribution():
    """Current tracker status distribution (for pie/doughnut chart)."""
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403

    pipeline = [
        {'$group': {'_id': '$status', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    results = list(mongo.db.trackers.aggregate(pipeline))

    status_labels = {
        STATUS_WAITING_NOC: 'Waiting NOC',
        STATUS_NOC_WORKING: 'NOC Working',
        STATUS_ZTP_PULL_PENDING: 'ZTP Pull Pending',
        STATUS_ZTP_CONFIG_UNVERIFIED: 'ZTP Config Unverified',
        STATUS_ZTP_PULL_DONE_FE: 'ZTP Pull Done (FE)',
        STATUS_ZTP_PULL_UNVERIFIED: 'ZTP Pull Unverified',
        STATUS_ZTP_PULL_REQ_NOC: 'ZTP Requested NOC',
        STATUS_FE_REQ_ZTP: 'FE Requested ZTP',
        STATUS_READY_COORD: 'Ready for Coordination',
        STATUS_HSO_SUBMITTED: 'HSO Submitted',
        STATUS_HSO_REJECTED: 'HSO Rejected',
        STATUS_COMPLETE: 'Completed',
    }
    colors = ['#EF4444','#F59E0B','#6366F1','#EC4899','#3B82F6','#8B5CF6',
              '#0EA5E9','#14B8A6','#22C55E','#84CC16','#F97316','#10B981']

    labels = [status_labels.get(r['_id'], r['_id']) for r in results]
    values = [r['count'] for r in results]

    return jsonify({
        'labels': labels,
        'datasets': [{
            'data': values,
            'backgroundColor': colors[:len(values)],
        }]
    })


@app.route('/api/analytics/ztp-breakdown')
@login_required
def api_analytics_ztp_breakdown():
    """ZTP performed by FE vs NS breakdown (for pie chart)."""
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    date_filter = {'created_at': {'$gte': start, '$lte': end}}

    by_fe = mongo.db.trackers.count_documents({**date_filter, 'ztp.performed_by': 'FE'})
    by_ns = mongo.db.trackers.count_documents({**date_filter, 'ztp.performed_by': 'NS'})
    pending = mongo.db.trackers.count_documents({**date_filter, 'ztp.performed_by': None, 'status': {'$ne': STATUS_COMPLETE}})

    return jsonify({
        'labels': ['ZTP by FE', 'ZTP by NOC', 'Pending'],
        'datasets': [{
            'data': [by_fe, by_ns, pending],
            'backgroundColor': ['#3B82F6', '#8B5CF6', '#D1D5DB'],
        }]
    })


@app.route('/api/analytics/sim-performance')
@login_required
def api_analytics_sim_performance():
    """SIM activation success vs failure rates."""
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    date_filter = {'created_at': {'$gte': start, '$lte': end}}

    sim1_total = mongo.db.trackers.count_documents({**date_filter, 'sim.sim1.status': {'$ne': 'not_required'}})
    sim1_ok = mongo.db.trackers.count_documents({**date_filter, 'sim.sim1.status': {'$in': ['activation_complete_manual', 'activation_complete_preactivated']}})
    sim1_fail = mongo.db.trackers.count_documents({**date_filter, 'sim.sim1.failure_reason': {'$ne': None}})
    sim1_pending = sim1_total - sim1_ok - sim1_fail

    sim2_total = mongo.db.trackers.count_documents({**date_filter, 'sim.sim2.status': {'$ne': 'not_required'}})
    sim2_ok = mongo.db.trackers.count_documents({**date_filter, 'sim.sim2.status': {'$in': ['activation_complete_manual', 'activation_complete_preactivated']}})
    sim2_fail = mongo.db.trackers.count_documents({**date_filter, 'sim.sim2.failure_reason': {'$ne': None}})
    sim2_pending = sim2_total - sim2_ok - sim2_fail

    return jsonify({
        'sim1': {'total': sim1_total, 'success': sim1_ok, 'failed': sim1_fail, 'pending': sim1_pending},
        'sim2': {'total': sim2_total, 'success': sim2_ok, 'failed': sim2_fail, 'pending': sim2_pending},
        'labels': ['SIM1 Success', 'SIM1 Failed', 'SIM1 Pending', 'SIM2 Success', 'SIM2 Failed', 'SIM2 Pending'],
        'datasets': [{
            'label': 'SIM Performance',
            'data': [sim1_ok, sim1_fail, sim1_pending, sim2_ok, sim2_fail, sim2_pending],
            'backgroundColor': ['#22C55E', '#EF4444', '#D1D5DB', '#10B981', '#F97316', '#E5E7EB'],
        }]
    })


@app.route('/api/analytics/sim-provider-performance')
@login_required
def api_analytics_sim_provider_performance():
    """SIM activation stats broken down by provider (Airtel, Jio, VI, BSNL)."""
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    range_type = request.args.get('range', 'today')
    start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
    date_filter = {'created_at': {'$gte': start, '$lte': end}}

    PROVIDERS = ['Airtel', 'Jio', 'VI', 'BSNL']
    SUCCESS_STATUSES = ['activation_complete_manual', 'activation_complete_preactivated']

    def agg_by_provider(sim_key):
        """Return {provider: {total, success, failed}} via a single aggregation."""
        pipeline = [
            {'$match': {**date_filter, f'sim.{sim_key}.status': {'$ne': 'not_required'},
                        f'sim.{sim_key}.provider': {'$in': PROVIDERS}}},
            {'$group': {
                '_id': f'$sim.{sim_key}.provider',
                'total': {'$sum': 1},
                'success': {'$sum': {'$cond': {
                    'if': {'$in': [f'$sim.{sim_key}.status', SUCCESS_STATUSES]},
                    'then': 1, 'else': 0
                }}},
                'failed': {'$sum': {'$cond': {
                    'if': {'$and': [
                        {'$ne': [f'$sim.{sim_key}.failure_reason', None]},
                        {'$ne': [f'$sim.{sim_key}.failure_reason', '']}
                    ]},
                    'then': 1, 'else': 0
                }}}
            }}
        ]
        return {row['_id']: row for row in mongo.db.trackers.aggregate(pipeline)}

    sim1_stats = agg_by_provider('sim1')
    sim2_stats = agg_by_provider('sim2')

    result = {}
    for p in PROVIDERS:
        s1 = sim1_stats.get(p, {'total': 0, 'success': 0, 'failed': 0})
        s2 = sim2_stats.get(p, {'total': 0, 'success': 0, 'failed': 0})
        total   = s1['total']   + s2['total']
        success = s1['success'] + s2['success']
        failed  = s1['failed']  + s2['failed']
        result[p] = {
            'total':   total,
            'success': success,
            'failed':  failed,
            'pending': max(0, total - success - failed),
        }

    return jsonify({'providers': PROVIDERS, 'data': result})


@app.route('/api/NOC_SUPPORT_GROUP/fe/day/<date>')
@app.route('/api/analytics/fe/day/<date>')
@login_required
def api_analytics_fe_day(date):
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        day_start = datetime.strptime(date, '%Y-%m-%d')
        day_end   = day_start + timedelta(days=1)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    trackers = list(mongo.db.trackers.find({'created_at': {'$gte': day_start, '$lt': day_end}}).sort('created_at', 1))
    timeline_data = []
    for t in trackers:
        noc_id   = t.get('noc_assignee')
        noc_name = None
        if noc_id:
            u = mongo.db.users.find_one({'_id': ObjectId(noc_id)})
            noc_name = u.get('name', u.get('username')) if u else None
        stage_times = calculate_stage_times(t)
        timeline_data.append({
            'tracker_id': t.get('tracker_id'),
            'sdwan_id':   t.get('sdwan_id'),
            'customer':   t.get('customer'),
            'fe_name':    t.get('fe', {}).get('name'),
            'fe_username': t.get('fe', {}).get('username'),
            'fe_phone':   t.get('fe', {}).get('phone'),
            'noc_assignee': noc_name,
            'status':      t.get('status'),
            'created_at':  t['created_at'].isoformat() if t.get('created_at') else None,
            'completed_at': t['completed_at'].isoformat() if t.get('completed_at') else None,
            'stage_times': stage_times,
            'events': [{'stage': e.get('stage'), 'timestamp': e['timestamp'].isoformat(), 'remarks': e.get('remarks')}
                       for e in t.get('events', []) if isinstance(e.get('timestamp'), datetime)]
        })
    return jsonify({'date': date, 'trackers': timeline_data})


@app.route('/api/NOC_SUPPORT_GROUP/noc/day/<date>')
@app.route('/api/analytics/noc/day/<date>')
@login_required
def api_analytics_noc_day(date):
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        day_start = datetime.strptime(date, '%Y-%m-%d')
        day_end   = day_start + timedelta(days=1)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    trackers = list(mongo.db.trackers.find({'created_at': {'$gte': day_start, '$lt': day_end}}))
    noc_data = {}
    for t in trackers:
        noc_id = t.get('noc_assignee')
        if noc_id:
            u = mongo.db.users.find_one({'_id': ObjectId(noc_id)})
            noc_name = u.get('name', u.get('username')) if u else 'Unknown'
        else:
            noc_id, noc_name = 'unassigned', 'Unassigned'
        key = f"{noc_id}|{noc_name}"
        noc_data.setdefault(key, {'noc_id': noc_id, 'noc_name': noc_name, 'started': 0, 'completed': 0})
        noc_data[key]['started'] += 1
        if t.get('completed_at'):
            noc_data[key]['completed'] += 1

    keys       = sorted(noc_data.keys())
    noc_labels = [noc_data[k]['noc_name']  for k in keys]
    noc_ids    = [noc_data[k]['noc_id']    for k in keys]
    return jsonify({
        'date': date,
        'labels': noc_labels,
        'noc_ids': noc_ids,
        'datasets': [
            {'label': 'Started',   'data': [noc_data[k]['started']   for k in keys], 'backgroundColor': '#3B82F6'},
            {'label': 'Completed', 'data': [noc_data[k]['completed'] for k in keys], 'backgroundColor': '#10B981'},
        ]
    })


@app.route('/api/NOC_SUPPORT_GROUP/noc/user/<user_id>/day/<date>')
@app.route('/api/analytics/noc/user/<user_id>/day/<date>')
@login_required
def api_analytics_noc_user_day(user_id, date):
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        day_start = datetime.strptime(date, '%Y-%m-%d')
        day_end   = day_start + timedelta(days=1)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    u = mongo.db.users.find_one({'_id': ObjectId(user_id)})
    noc_name = u.get('name', u.get('username')) if u else 'Unknown'
    trackers = list(mongo.db.trackers.find({'created_at': {'$gte': day_start, '$lt': day_end},
                                            'noc_assignee': user_id}).sort('created_at', 1))
    timeline_data = []
    for t in trackers:
        st = calculate_stage_times(t)
        timeline_data.append({
            'tracker_id': t.get('tracker_id'),
            'sdwan_id':   t.get('sdwan_id'),
            'customer':   t.get('customer'),
            'fe_name':    t.get('fe', {}).get('name'),
            'status':     t.get('status'),
            'created_at': t['created_at'].isoformat() if t.get('created_at') else None,
            'completed_at': t['completed_at'].isoformat() if t.get('completed_at') else None,
            'stage_times': st,
            'events': [{'stage': e.get('stage'), 'timestamp': e['timestamp'].isoformat(), 'remarks': e.get('remarks')}
                       for e in t.get('events', []) if isinstance(e.get('timestamp'), datetime)]
        })
    return jsonify({'date': date, 'noc_name': noc_name, 'trackers': timeline_data})


@app.route('/api/NOC_SUPPORT_GROUP/export/fe')
@app.route('/api/analytics/export/fe')
@login_required
def api_analytics_export_fe():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        range_type = request.args.get('range', 'today')
        start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
        trackers   = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}).sort('created_at', 1))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FE Analytics"

        headers = ['SDWAN ID','Customer','FE Name','FE Username','FE Phone','NOC Assignee',
                   'Status','Created At','Completed At','Duration (hours)',
                   'Queue Wait (min)','SIM1 Act (min)','NS Processing (min)','ZTP Exec (min)',
                   'ZTP Performed By','HSO Attempts','SIM1 Failed','ZTP Failed']
        ws.append(headers)
        hfill = PatternFill(start_color='1A56A0', end_color='1A56A0', fill_type='solid')
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        for t in trackers:
            created   = t.get('created_at')
            completed = t.get('completed_at')
            duration  = round((completed - created).total_seconds() / 3600, 2) if created and completed else ''
            noc_id    = t.get('noc_assignee')
            noc_name  = ''
            if noc_id:
                u = mongo.db.users.find_one({'_id': ObjectId(noc_id)})
                noc_name = u.get('name', u.get('username', '')) if u else ''
            st  = calculate_stage_times(t)
            ztp = t.get('ztp', {})
            hso = t.get('hso', {})
            ws.append([
                t.get('sdwan_id',''), t.get('customer',''),
                t.get('fe',{}).get('name',''), t.get('fe',{}).get('username',''), t.get('fe',{}).get('phone',''),
                noc_name, t.get('status',''),
                created.strftime('%Y-%m-%d %H:%M:%S') if created else '',
                completed.strftime('%Y-%m-%d %H:%M:%S') if completed else '',
                duration,
                round(st['queue_wait_minutes'], 1) if st.get('queue_wait_minutes') is not None else '',
                round(st['sim1_activation_minutes'], 1) if st.get('sim1_activation_minutes') is not None else '',
                round(st['ns_processing_minutes'], 1) if st.get('ns_processing_minutes') is not None else '',
                round(st['ztp_execution_minutes'], 1) if st.get('ztp_execution_minutes') is not None else '',
                ztp.get('performed_by', ''),
                len(hso.get('attempts', [])),
                'Yes' if t.get('sim',{}).get('sim1',{}).get('failure_reason') else 'No',
                'Yes' if ztp.get('failure_reason') else 'No',
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col) + 2, 50)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'fe_analytics_{range_type}_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/NOC_SUPPORT_GROUP/export/noc')
@app.route('/api/analytics/export/noc')
@login_required
def api_analytics_export_noc():
    if not _analytics_allowed():
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        range_type = request.args.get('range', 'today')
        start, end = get_date_range(range_type, request.args.get('from'), request.args.get('to'))
        trackers   = list(mongo.db.trackers.find({'created_at': {'$gte': start, '$lte': end}}).sort('created_at', 1))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NOC Analytics"

        headers = ['SDWAN ID','Customer','FE Name','NOC Assignee','Status',
                   'Created At','NOC Assigned At','Completed At',
                   'Queue Wait (min)','NS Processing (min)','ZTP Performed By','Total Duration (h)']
        ws.append(headers)
        hfill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        for t in trackers:
            created   = t.get('created_at')
            completed = t.get('completed_at')
            noc_id    = t.get('noc_assignee')
            noc_name  = ''
            if noc_id:
                u = mongo.db.users.find_one({'_id': ObjectId(noc_id)})
                noc_name = u.get('name', u.get('username', '')) if u else ''
            st = calculate_stage_times(t)
            assigned_at = t.get('stage_timestamps', {}).get('noc_assigned_at')
            ws.append([
                t.get('sdwan_id',''), t.get('customer',''),
                t.get('fe',{}).get('name',''), noc_name, t.get('status',''),
                created.strftime('%Y-%m-%d %H:%M:%S') if created else '',
                assigned_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(assigned_at, datetime) else '',
                completed.strftime('%Y-%m-%d %H:%M:%S') if completed else '',
                round(st['queue_wait_minutes'], 1) if st.get('queue_wait_minutes') is not None else '',
                round(st['ns_processing_minutes'], 1) if st.get('ns_processing_minutes') is not None else '',
                t.get('ztp',{}).get('performed_by',''),
                round((completed - created).total_seconds()/3600, 2) if created and completed else '',
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col) + 2, 50)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'noc_analytics_{range_type}_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Admin: User Management ─────────────────────────────────────────────────
# Separate password-gated admin panel for managing users (create / update).
# Admin session is tracked independently of the regular user session.

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'qwerty')

# Role → human label mapping for display
ROLE_LABELS = {
    ROLE_FE:        'Field Engineer',
    ROLE_FEG:       'Field Engineer Group',
    ROLE_FS:        'Field Support',
    ROLE_FSG:       'Field Support Group',
    ROLE_NS:        'NOC Support',
    ROLE_NSG:       'NOC Support Group',
    ROLE_ANALYTICS: 'Analytics',
}

# Fields that are relevant for each role (used by the frontend to show/hide inputs)
ROLE_FIELDS = {
    ROLE_FE:  ['name', 'username', 'password', 'zone', 'region', 'state',
               'field_engineer_group', 'field_support', 'email', 'contact', 'location'],
    ROLE_FEG: ['name', 'username', 'password', 'zone', 'region', 'state', 'field_support'],
    ROLE_FS:  ['name', 'username', 'password', 'zone', 'region', 'field_support_group'],
    ROLE_FSG: ['name', 'username', 'password', 'zone'],
    ROLE_NS:  ['name', 'username', 'password'],
    ROLE_NSG: ['name', 'username', 'password'],
    ROLE_ANALYTICS: ['name', 'username', 'password'],
}


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_authenticated'):
            return jsonify({'error': 'Admin authentication required'}), 403
        return f(*args, **kwargs)
    return decorated


@app.route('/admin')
def admin_page():
    return render_template('admin_users.html',
                           role_labels=ROLE_LABELS,
                           role_fields=ROLE_FIELDS,
                           all_roles=list(ROLE_LABELS.keys()))


@app.route('/admin/auth', methods=['POST'])
def admin_auth():
    data = request.json or {}
    if data.get('password') == ADMIN_PASSWORD:
        session['admin_authenticated'] = True
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Incorrect password'}), 401


@app.route('/admin/logout', methods=['POST'])
def admin_logout():
    session.pop('admin_authenticated', None)
    return jsonify({'success': True})


@app.route('/admin/api/users', methods=['GET'])
@admin_required
def admin_list_users():
    users = list(mongo.db.users.find({}, {'password': 0}))
    for u in users:
        u['_id'] = str(u['_id'])
    return jsonify(users)


@app.route('/admin/api/users', methods=['POST'])
@admin_required
def admin_create_user():
    data = request.json or {}
    role = data.get('role')
    if role not in ROLE_LABELS:
        return jsonify({'error': 'Invalid role'}), 400

    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')

    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if not password:
        return jsonify({'error': 'Password is required'}), 400

    # username defaults to name if not provided
    if not username:
        username = name

    # Prevent duplicate usernames
    if mongo.db.users.find_one({'username': username}):
        return jsonify({'error': f'Username "{username}" already exists'}), 409

    doc = {
        'name':     name,
        'username': username,
        'password': generate_password_hash(password),
        'role':     role,
        'created_at': get_utc_now(),
    }
    # Optional role-specific fields
    for field in ['zone', 'region', 'state', 'field_engineer_group', 'field_support',
                  'field_support_group', 'email', 'contact', 'location']:
        val = (data.get(field) or '').strip()
        if val:
            doc[field] = val

    result = mongo.db.users.insert_one(doc)
    return jsonify({'success': True, 'id': str(result.inserted_id)}), 201


@app.route('/admin/api/users/<user_id>', methods=['PUT'])
@admin_required
def admin_update_user(user_id):
    data = request.json or {}
    try:
        oid = ObjectId(user_id)
    except Exception:
        return jsonify({'error': 'Invalid user ID'}), 400

    user = mongo.db.users.find_one({'_id': oid})
    if not user:
        return jsonify({'error': 'User not found'}), 404

    set_ops = {}

    # Name
    name = (data.get('name') or '').strip()
    if name:
        set_ops['name'] = name

    # Username — check uniqueness if changed
    username = (data.get('username') or '').strip()
    if username and username != user.get('username'):
        if mongo.db.users.find_one({'username': username, '_id': {'$ne': oid}}):
            return jsonify({'error': f'Username "{username}" already exists'}), 409
        set_ops['username'] = username

    # Password — only update if provided
    password = data.get('password', '')
    if password:
        set_ops['password'] = generate_password_hash(password)

    # Optional fields
    for field in ['zone', 'region', 'state', 'field_engineer_group', 'field_support',
                  'field_support_group', 'email', 'contact', 'location']:
        if field in data:
            set_ops[field] = (data[field] or '').strip()

    if not set_ops:
        return jsonify({'error': 'No changes provided'}), 400

    set_ops['updated_at'] = get_utc_now()
    mongo.db.users.update_one({'_id': oid}, {'$set': set_ops})
    return jsonify({'success': True})


# ─── Socket.IO Event Handlers ───────────────────────────────────────────────
# These handlers manage real-time WebSocket connections for instant updates

@socketio.on('connect')
def handle_connect():
    """Client connected to WebSocket"""
    print(f"Client connected: {request.sid}")

@socketio.on('disconnect')
def handle_disconnect():
    """Client disconnected from WebSocket"""
    print(f"Client disconnected: {request.sid}")

@socketio.on('join_tracker')
def handle_join_tracker(data):
    """User joins a tracker room to receive real-time updates for that tracker"""
    tracker_id = data.get('tracker_id')
    if tracker_id:
        join_room(f"tracker_{tracker_id}")
        print(f"Client {request.sid} joined tracker_{tracker_id}")

@socketio.on('leave_tracker')
def handle_leave_tracker(data):
    """User leaves a tracker room"""
    tracker_id = data.get('tracker_id')
    if tracker_id:
        leave_room(f"tracker_{tracker_id}")
        print(f"Client {request.sid} left tracker_{tracker_id}")

@socketio.on('join_dashboard')
def handle_join_dashboard(data):
    """User joins their dashboard room to receive tracker list updates"""
    user_id = data.get('user_id')
    role = data.get('role')
    if user_id and role:
        # Join role-specific room for dashboard updates
        join_room(f"dashboard_{role}")
        join_room(f"user_{user_id}")
        print(f"Client {request.sid} joined dashboard_{role} and user_{user_id}")

@socketio.on('leave_dashboard')
def handle_leave_dashboard(data):
    """User leaves dashboard room"""
    user_id = data.get('user_id')
    role = data.get('role')
    if user_id and role:
        leave_room(f"dashboard_{role}")
        leave_room(f"user_{user_id}")
        print(f"Client {request.sid} left dashboard_{role} and user_{user_id}")


# ─── Helper Functions for Broadcasting ──────────────────────────────────────
def broadcast_tracker_update(tracker_id, event_type, data, include_full_tracker=True):
    """
    Broadcast tracker updates with optional complete tracker data
    
    Args:
        tracker_id: Tracker ID
        event_type: Type of event (e.g., 'ztp_status_updated')
        data: Partial data (for backward compatibility)
        include_full_tracker: If True, fetch and include complete tracker
    """
    payload = {
        'tracker_id': tracker_id,
        'event_type': event_type,
        'data': data,
        'timestamp': datetime.utcnow().isoformat() + 'Z'
    }
    
    # Include complete tracker data if enabled and mode supports it
    if include_full_tracker and SOCKET_INCLUDE_FULL_DATA and REALTIME_MODE in ('socket', 'hybrid'):
        try:
            tracker = mongo.db.trackers.find_one({'_id': ObjectId(tracker_id)})
            if tracker:
                payload['tracker'] = serialize_doc(tracker)
                print(f"[Socket.IO] Broadcasting tracker_update with full data: tracker_id={tracker_id}, event_type={event_type}")
            else:
                print(f"[Socket.IO] Broadcasting tracker_update (tracker not found): tracker_id={tracker_id}, event_type={event_type}")
        except Exception as e:
            print(f"[Socket.IO] Error fetching tracker for broadcast: {e}")
            print(f"[Socket.IO] Broadcasting tracker_update without full data: tracker_id={tracker_id}, event_type={event_type}")
    else:
        print(f"[Socket.IO] Broadcasting tracker_update: tracker_id={tracker_id}, event_type={event_type}")
    
    socketio.emit('tracker_update', payload, room=f"tracker_{tracker_id}")

def broadcast_chat_message(tracker_id, message):
    """Broadcast new chat message to all users in tracker room"""
    print(f"[Socket.IO] Broadcasting new_chat_message: tracker_id={tracker_id}")
    socketio.emit('new_chat_message', {
        'tracker_id': tracker_id,
        'message': serialize_doc(message)
    }, room=f"tracker_{tracker_id}")

def broadcast_dashboard_update(role, event_type, data):
    """Broadcast dashboard updates to all users of a specific role"""
    print(f"[Socket.IO] Broadcasting dashboard_update: role={role}, event_type={event_type}")
    socketio.emit('dashboard_update', {
        'event_type': event_type,
        'data': data
    }, room=f"dashboard_{role}")

def broadcast_to_user(user_id, event_type, data):
    """Broadcast notification to a specific user"""
    socketio.emit('user_notification', {
        'event_type': event_type,
        'data': data
    }, room=f"user_{user_id}")


if __name__ == '__main__':
    # eventlet.monkey_patch() at module top gives Flask-SocketIO native WebSocket
    # support. socketio.run() uses eventlet's WSGI server — no Waitress/Gunicorn needed.
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    port = int(os.environ.get('PORT', 5001))
    socketio.run(app, debug=debug, host='0.0.0.0', port=port)

