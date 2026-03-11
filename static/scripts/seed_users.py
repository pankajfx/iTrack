"""
seed_users.py
─────────────
Reads user data from the Excel file:
    static/scripts/SDWAN Installation Tracker Master User Data.xlsx

Clears the users collection and re-populates it with all users.
Passwords are hashed with Werkzeug. FEG hierarchy (field_support) is
derived from region → FS username mapping.

Usage (run from project root):
    python static/scripts/seed_users.py

Requirements:
    pip install openpyxl werkzeug pymongo
"""

import os
import sys
from collections import Counter
from datetime import datetime

from bson import ObjectId
from pymongo import MongoClient
from werkzeug.security import generate_password_hash

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
MONGO_URI = os.environ.get('MONGO_URI', 'mongodb://localhost:27017/sdwan_tracker')
#prod
# MONGO_URI  = os.environ.get('MONGO_URI', 'mongodb://myAdminUser:MyStrongPassword123@localhost:27017/sdwan_tracker?authSource=admin')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'SDWAN Installation Tracker Master User Data.xlsx')

# ── Role mapping: Excel label → app constant ──────────────────────────────────
ROLE_MAP = {
    'Field Support Group':  'FIELD_SUPPORT_GROUP',
    'Field Support':        'FIELD_SUPPORT',
    'Field Engineer Group': 'FIELD_ENGINEER_GROUP',
    'Field Engineer':       'FIELD_ENGINEER',
    'NOC Support Group':    'NOC_SUPPORT_GROUP',
    'NOC Support':          'NOC_SUPPORT',
    'Analytics':            'ANALYTICS',
}

# ── Read Excel ────────────────────────────────────────────────────────────────
def read_excel_rows():
    """Return list of dicts from the 'All(270)' sheet (header row → keys)."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    # Use the consolidated sheet; fall back to first sheet if name changes
    sheet_name = 'All(270)' if 'All(270)' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if row.get('Role'):          # skip blank trailing rows
            rows.append(row)
    return rows


# ── Build MongoDB documents ───────────────────────────────────────────────────
def build_user_docs(rows):
    now = datetime.utcnow()

    # Build region → FS username lookup (used to set field_support on FEG docs)
    region_to_fs = {}
    for row in rows:
        if row.get('Role') == 'Field Support':
            region = row.get('Region')
            if region:
                region_to_fs[region] = str(row['Username']).strip()

    docs = []
    skipped = []

    for row in rows:
        excel_role = str(row.get('Role', '')).strip()
        role = ROLE_MAP.get(excel_role)
        if not role:
            skipped.append(f"Unknown role '{excel_role}' — username: {row.get('Username')}")
            continue

        username       = str(row.get('Username', '')).strip()
        name           = str(row.get('Name', username)).strip()
        password_plain = str(row.get('Password', username)).strip()
        zone           = str(row['Zone']).strip() if row.get('Zone') else 'India'
        region         = str(row['Region']).strip() if row.get('Region') else None
        group          = str(row['Group']).strip() if row.get('Group') else None   # FE → FEG name
        state          = str(row['State']).strip() if row.get('State') else None   # FE → location
        email          = str(row['Email']).strip() if row.get('Email') else None
        # Contact may be stored as float (e.g. 6302828144.0) — normalise to string
        raw_contact    = row.get('Contact')
        contact        = str(int(raw_contact)) if isinstance(raw_contact, float) else (
                         str(raw_contact).strip() if raw_contact else None)

        doc = {
            'username':   username,
            'name':       name,
            'password':   generate_password_hash(password_plain),
            'role':       role,
            'active':     True,
            'created_at': now,
            'updated_at': now,
            'zone':       zone,
        }

        if region:
            doc['region'] = region

        # Role-specific hierarchy / contact fields
        if role == 'FIELD_ENGINEER':
            if group:
                doc['field_engineer_group'] = group
            if state:
                doc['location'] = state
            if email:
                doc['email'] = email
            if contact:
                doc['contact'] = contact

        elif role == 'FIELD_ENGINEER_GROUP':
            # Derive parent FS from region
            if region and region in region_to_fs:
                doc['field_support'] = region_to_fs[region]

        docs.append(doc)

    if skipped:
        print(f"\n  WARNINGS ({len(skipped)}):")
        for msg in skipped:
            print(f"    {msg}")

    return docs


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Excel source: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print("ERROR: Excel file not found.")
        sys.exit(1)

    rows = read_excel_rows()
    print(f"  Read {len(rows)} rows from Excel")

    docs = build_user_docs(rows)
    print(f"  Built {len(docs)} user documents")

    client = MongoClient(MONGO_URI)
    db     = client.get_default_database()

    # Ensure collection + unique index on username
    if 'users' not in db.list_collection_names():
        db.create_collection('users')
        print("  Created 'users' collection")
    db.users.create_index('username', unique=True)

    print("\nClearing existing users...")
    deleted = db.users.delete_many({}).deleted_count
    print(f"  Deleted {deleted} existing user(s)")

    print(f"Inserting {len(docs)} users...")
    result = db.users.insert_many(docs)
    print(f"  Inserted {len(result.inserted_ids)} user(s)")

    # Summary
    counts = Counter(d['role'] for d in docs)
    print("\nUsers by role:")
    for role, count in sorted(counts.items()):
        print(f"  {role:<30} {count}")
    print(f"\nTotal: {len(docs)} users seeded successfully.")

    client.close()


if __name__ == '__main__':
    main()
